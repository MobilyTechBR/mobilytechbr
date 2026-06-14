from __future__ import annotations

import csv
import html
import json
import math
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFilter, ImageFont, ImageOps


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"
CREATIVE_DIR = ROOT / "assets" / "phase2-creatives"
BRAND_LOGO_DIR = ROOT / "assets" / "brand-logos"
BRAND_OFFICIAL_DIR = ROOT / "assets" / "brand-officials"
AFFILIATE_IMAGE_DIR = ROOT / "assets" / "phase2-affiliates"
OUTPUT_DIR = Path.home() / "Documents" / "New project" / "outputs" / "mobilytech_fase2_hibrida_2026-06-13"

GENERATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
BACKUP_PATH = Path.home() / "Documents" / "Site Vercel backups" / "Site Vercel backup antes fase2 hibrida 2026-06-13_08-48-34.zip"
WIX_PREMIUM_SITE_ID = "85e985c5-2904-452f-85e2-a98f6d3b1cac"
WIX_DOMAIN = "https://www.mobilytech.com.br/"
VERCEL_PRODUCTION = "https://mobilytechbr.vercel.app"

BRANDS = [
    {"id": "intel", "name": "Intel", "accent": "#38bdf8"},
    {"id": "amd", "name": "AMD", "accent": "#22c55e"},
    {"id": "nvidia", "name": "NVIDIA", "accent": "#76b900"},
    {"id": "microsoft", "name": "Microsoft", "accent": "#60a5fa"},
    {"id": "corsair", "name": "Corsair", "accent": "#f8fafc"},
    {"id": "asus", "name": "ASUS", "accent": "#9ca3af"},
    {"id": "gigabyte", "name": "Gigabyte", "accent": "#22d3ee"},
    {"id": "evga", "name": "EVGA", "accent": "#d1d5db"},
    {"id": "kingston", "name": "Kingston", "accent": "#ef4444"},
    {"id": "crucial", "name": "Crucial", "accent": "#60a5fa"},
    {"id": "pny", "name": "PNY", "accent": "#94a3b8"},
]

SALES_COPY = {
    "case-ssd-nvme": "Transforme um SSD NVMe parado em armazenamento externo rapido para jogos, backups e arquivos pesados. Um upgrade pequeno que combina muito com quem compra PC revisado e quer mais espaco sem abrir o gabinete.",
    "case-ssd-sata": "Reaproveite HDs e SSDs de notebook como backup externo em poucos minutos. E barato, util e resolve aquela falta de espaco sem complicar o setup.",
    "fone-kz-castor": "Som limpo, grave forte e visual discreto para jogar, estudar e ouvir musica sem depender de headset grande. Uma escolha custo-beneficio para setup gamer compacto.",
    "kit-limpeza-esd": "Kit pratico para tirar poeira, cuidar dos contatos e manter o PC com cara de novo entre uma limpeza profissional e outra. Ideal para quem quer preservar desempenho e aparencia.",
    "mini-aspirador-teclado": "Para mesa, teclado, notebook e cantos do setup que acumulam poeira todo dia. Um acessorio simples para manter o ambiente mais limpo sem desmontar nada.",
    "hub-usb-c": "Faltou porta USB no notebook ou no setup? Esse hub resolve mouse, teclado, pendrive e perifericos em uma conexao so, sem gambiarra.",
    "suporte-gpu-antisag": "Ajuda a deixar a placa de video reta, melhora o visual interno do gabinete e reduz tensao no slot PCIe. Pequeno detalhe que deixa o setup mais profissional.",
    "keycaps-pbt": "Troque o visual do teclado mecanico sem comprar outro teclado. Um upgrade barato para personalizar o setup e dar cara nova para a mesa.",
    "teclado-mecanico-abnt2": "Teclado mecanico com RGB e layout ABNT2 para jogar, digitar e montar um setup completo. Bom para quem quer sentir o upgrade todo dia.",
    "bias-light-led": "Luz de fundo para monitor que deixa o setup mais bonito e ajuda no conforto visual em ambientes escuros. Um detalhe barato que muda a aparencia da mesa.",
}

NAV_ICONS = {
    "home": '<path d="M4 11.5 12 5l8 6.5V20h-5v-5H9v5H4z"/>',
    "ofertas": '<path d="M4 7h16v4H4z"/><path d="M6 11h12v8H6z"/><path d="M9 7V5h6v2"/>',
    "montagem": '<path d="M5 8h14v8H5z"/><path d="M9 20h6"/><path d="M12 16v4"/><path d="M8 11h2m4 0h2"/>',
    "limpeza": '<path d="M6 14c3-1 4-4 4-9 3 2 5 5 6 9"/><path d="M5 14h14l-1 6H6z"/><path d="M8 17h8"/>',
    "achados": '<path d="m12 3 1.8 5.4 5.2 1.8-5.2 1.8L12 17.5 10.2 12 5 10.2l5.2-1.8z"/><path d="m18 15 .7 2.1 2.1.7-2.1.7-.7 2.1-.7-2.1-2.1-.7 2.1-.7z"/>',
    "avaliacoes": '<path d="M12 3 9.2 8.8 3 9.7l4.5 4.4-1.1 6.2L12 17.4l5.6 2.9-1.1-6.2L21 9.7l-6.2-.9z"/>',
    "contato": '<path d="M4 6h16v12H4z"/><path d="m4 7 8 6 8-6"/>',
}


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def trends_link(query: str) -> str:
    return f"https://trends.google.com/trends/explore?geo=BR&q={quote(query)}"


def meta_link(query: str) -> str:
    return (
        "https://www.facebook.com/ads/library/"
        f"?active_status=active&ad_type=all&country=BR&q={quote(query)}"
        "&search_type=keyword_unordered&media_type=all"
    )


def ml_trends_link(query: str) -> str:
    return f"https://tendencias.mercadolivre.com.br/{quote(query.replace(' ', '-'))}"


def ml_search_link(query: str) -> str:
    return f"https://lista.mercadolivre.com.br/{quote(query.replace(' ', '-'))}"


FINALISTS = [
    {
        "id": "case-ssd-nvme",
        "title": "Case SSD M.2 NVMe USB-C 10Gbps",
        "niche": "Armazenamento e upgrades",
        "platform": "Mercado Livre / Shopee",
        "sourceUrl": "https://www.mercadolivre.com.br/case-para-ssd-m2-nvme-externo-usb-c-31-adaptador-nvme-2230-2242-2260-2280-aluminio-m-key-ate-4tb-10gbps-inv-tech/p/MLB29349887",
        "currentPrice": "R$59,98 no Mercado Livre; Shopee em torno de R$74,57",
        "shipping": "Frete gratis/por cupom em ofertas vistas; prazo depende do vendedor e CEP",
        "delivery": "ML local tende a ser mais rapido; Shopee pode variar por vendedor",
        "sellerReputation": "+5 mil vendidos, 4.8/5 com 463 opinioes, 'Mais vendido' em gavetas para HDs na oferta Inv Tech",
        "reviews": "Boa prova social para velocidade, praticidade e uso plug-and-play",
        "returnPolicy": "Mercado Livre indica Compra Garantida e devolucao gratis quando aplicavel; confirmar oferta final antes de anunciar",
        "operationModel": "Testar afiliado primeiro e depois dropshipping automatizado",
        "whySell": "Casa muito bem com SSDs, upgrades e clientes que compram PCs usados/revisados.",
        "costTarget": "R$58 a R$75",
        "sellTarget": "R$119 a R$139",
        "margin": "Margem bruta alvo de R$40 a R$65 antes de taxas",
        "confidence": "Alta",
        "wixCategory": "MobilyTech Finds / Armazenamento",
        "creativeAngles": [
            "Transforme SSD parado em armazenamento externo rapido",
            "Backup de jogos, fotos e projetos sem abrir o PC"
        ],
        "risk": "Precisa separar NVMe de SATA para evitar compra errada.",
    },
    {
        "id": "case-ssd-sata",
        "title": "Case SSD/HD SATA 2.5 USB 3.0",
        "niche": "Armazenamento e manutencao",
        "platform": "Mercado Livre",
        "sourceUrl": "https://www.mercadolivre.com.br/case-slim-hd-ssd-adaptador-usb-30-sata-3-externo-6gbps-ps4/p/MLB27786386",
        "currentPrice": "R$22 com opcoes a partir de R$17,50",
        "shipping": "Algumas opcoes com frete gratis; envio nacional",
        "delivery": "Depende do vendedor; produto local facilita entrega rapida",
        "sellerReputation": "+10 mil vendidos, 4.8/5 com 906 opinioes; vendedor MercadoLider em oferta vista",
        "reviews": "Alta prova social: facil instalacao e bom custo-beneficio",
        "returnPolicy": "Devolucao gratis em 30 dias na oferta vista e Compra Garantida",
        "operationModel": "Afiliado primeiro, dropshipping automatizado se houver fornecedor estavel",
        "whySell": "Produto barato, demonstravel em video e perfeito para reaproveitar HD/SSD de notebook.",
        "costTarget": "R$17,50 a R$30",
        "sellTarget": "R$49 a R$69",
        "margin": "Margem bruta alvo de R$20 a R$35 antes de taxas",
        "confidence": "Alta",
        "wixCategory": "MobilyTech Finds / Armazenamento",
        "creativeAngles": [
            "Nao jogue fora o HD antigo: transforme em backup externo",
            "Upgrade simples para quem comprou PC revisado"
        ],
        "risk": "Produto comoditizado; diferencial precisa ser conteudo e kit com SSD/limpeza.",
    },
    {
        "id": "fone-kz-castor",
        "title": "Fone KZ Castor monitor in-ear",
        "niche": "Audio gamer e setup",
        "platform": "Mercado Livre / Shopee / AliExpress",
        "sourceUrl": "https://www.mercadolivre.com.br/fone-kz-castor-audio-monitor-de-palco-alta-fidelidade-cor-cinza-sem-microfone/p/MLB28114243",
        "currentPrice": "R$114 no Mercado Livre; Shopee oficial foi visto em torno de R$109 com cupom",
        "shipping": "ML com frete gratis em oferta vista; Shopee pode depender de cupom/importacao",
        "delivery": "ML mais rapido; Shopee/AliExpress podem ter prazo maior",
        "sellerReputation": "Produto com 4.8/5 e 175 opinioes; uma opcao alternativa ML tinha vendedor com +5 mil vendas",
        "reviews": "Comentarios destacam graves, clareza, custo-beneficio e uso para jogos/musica",
        "returnPolicy": "ML com devolucao gratis/Compra Garantida quando aplicavel; confirmar oferta final",
        "operationModel": "Afiliado primeiro; dropshipping so com fornecedor oficial/confiavel",
        "whySell": "Produto quente em TikTok/audio e encaixa bem com setup gamer sem parecer aleatorio no site.",
        "costTarget": "R$95 a R$115",
        "sellTarget": "R$149 a R$179",
        "margin": "Margem apertada no afiliado; melhor para conteudo e ticket complementar",
        "confidence": "Media-alta",
        "wixCategory": "MobilyTech Finds / Audio",
        "creativeAngles": [
            "Som de monitor gastando menos que headset gamer comum",
            "Setup limpo para jogar, estudar e ouvir musica"
        ],
        "risk": "Audio e muito subjetivo; evitar promessa exagerada de qualidade profissional.",
    },
    {
        "id": "kit-limpeza-esd",
        "title": "Kit limpeza ESD antiestatico para PC",
        "niche": "Limpeza e manutencao",
        "platform": "Mercado Livre / Shopee",
        "sourceUrl": "https://www.mercadolivre.com.br/kit-limpeza-esd-antiestatico-para-pc-com-5-pecas/p/MLB37963040",
        "currentPrice": "Ofertas entre cerca de R$21 e R$58 conforme composicao",
        "shipping": "Algumas ofertas com frete gratis; verificar cupom/CEP",
        "delivery": "Varia por vendedor; kits nacionais tendem a chegar antes",
        "sellerReputation": "Oferta vista de loja Tapcamp com +100 mil vendas",
        "reviews": "Sinal positivo por ser acessorio de baixo risco e uso recorrente",
        "returnPolicy": "Devolucao gratis em 30 dias quando indicada no ML",
        "operationModel": "Dropshipping automatizado com estoque reserva local para combos",
        "whySell": "Conecta diretamente com a secao de limpeza de PCs e pode virar upsell no checkout.",
        "costTarget": "R$21 a R$35",
        "sellTarget": "R$59 a R$89",
        "margin": "Boa margem em kit/combo; campanha fria precisa criativo forte",
        "confidence": "Alta",
        "wixCategory": "MobilyTech Finds / Limpeza",
        "creativeAngles": [
            "Poeira e estatica: o erro silencioso que mata desempenho",
            "Kit simples para manter teclado, placa e gabinete apresentaveis"
        ],
        "risk": "Nao prometer recuperacao de componente; vender como prevencao e cuidado.",
    },
    {
        "id": "mini-aspirador-teclado",
        "title": "Mini aspirador USB para teclado e cantos do setup",
        "niche": "Limpeza e setup",
        "platform": "Mercado Livre / Shopee",
        "sourceUrl": "https://produto.mercadolivre.com.br/MLB-3409651937-mini-aspirador-de-po-usb-computador-teclado-house-tools-_JM",
        "currentPrice": "R$21,62 na oferta House Tools",
        "shipping": "Envio para todo o pais; frete/prazo por CEP",
        "delivery": "Produto local com vendedor grande tende a ser operacionalmente simples",
        "sellerReputation": "Loja oficial Costa Atacado, MercadoLider Platinum, +100 mil vendas",
        "reviews": "Poucas avaliacoes; review indica que e fraco, mas aceitavel pelo preco",
        "returnPolicy": "Devolucao gratis indicada na oferta vista",
        "operationModel": "Afiliado primeiro; dropshipping apenas se qualidade for validada",
        "whySell": "Criativo de antes/depois funciona bem, mas precisa expectativa honesta.",
        "costTarget": "R$21 a R$30",
        "sellTarget": "R$49 a R$69",
        "margin": "Margem boa, mas risco de suporte maior se expectativa for mal comunicada",
        "confidence": "Media",
        "wixCategory": "MobilyTech Finds / Limpeza",
        "creativeAngles": [
            "Teclado cheio de migalha e poeira em 30 segundos",
            "Limpeza rapida para mesa, notebook e setup gamer"
        ],
        "risk": "Potencia limitada; criativo precisa ser honesto para reduzir reclamacao.",
    },
    {
        "id": "hub-usb-c",
        "title": "Hub USB-C 4 portas para notebook e celular",
        "niche": "Conectividade",
        "platform": "Mercado Livre / Amazon",
        "sourceUrl": "https://www.mercadolivre.com.br/hub-usb-type-c-5-gbps-extensor-adaptador-4-portas-usb-30-computador-pc-notebook-celular-smartphone-tablet-chrome-technology/p/MLB54987539",
        "currentPrice": "R$19,27 em oferta Chrome Technology vista em busca",
        "shipping": "Confirmar CEP; produto leve e facil de enviar",
        "delivery": "Boa opcao para afiliado por marketplaces locais",
        "sellerReputation": "Loja oficial Chrome Technology em oferta vista",
        "reviews": "Produto de necessidade recorrente; validar avaliacoes no link final",
        "returnPolicy": "Compra Garantida/marketplace; confirmar na oferta final",
        "operationModel": "Afiliado primeiro",
        "whySell": "Resolve problema comum de notebook com poucas portas e complementa PCs/notebooks.",
        "costTarget": "R$19 a R$35",
        "sellTarget": "R$49 a R$79",
        "margin": "Margem razoavel, melhor como produto de volume/remarketing",
        "confidence": "Media-alta",
        "wixCategory": "MobilyTech Finds / Conectividade",
        "creativeAngles": [
            "Falta porta USB no notebook? Resolva sem gambiarra",
            "Mouse, teclado, pendrive e headset no mesmo hub"
        ],
        "risk": "Concorrencia alta; diferenciar por curadoria e garantia do marketplace.",
    },
    {
        "id": "suporte-gpu-antisag",
        "title": "Suporte anti-sag para placa de video",
        "niche": "PC gamer e estetica",
        "platform": "Mercado Livre / AliExpress",
        "sourceUrl": "https://www.mercadolivre.com.br/suporte-de-gpu-anti-sag-com-im-e-almofada-de-borracha/p/MLB2028023362",
        "currentPrice": "Ofertas vistas de R$19,90 a R$134,99 conforme marca/RGB",
        "shipping": "Frete geralmente simples; confirmar vendedor final",
        "delivery": "Melhor testar fornecedor nacional ou afiliado antes",
        "sellerReputation": "Algumas ofertas tem poucas avaliacoes; precisa escolha cuidadosa",
        "reviews": "Sinal bom para setups com GPUs pesadas, mas volume menor que cases/hubs",
        "returnPolicy": "Confirmar devolucao e Compra Garantida na oferta escolhida",
        "operationModel": "Testar afiliado primeiro e so depois dropshipping",
        "whySell": "Visual forte para criativos e conversa com PCs gamer montados.",
        "costTarget": "R$19 a R$45 para versao simples; RGB pode passar de R$100",
        "sellTarget": "R$59 a R$99 simples; R$129+ RGB",
        "margin": "Boa margem se fornecedor barato e qualidade aceitavel",
        "confidence": "Media",
        "wixCategory": "MobilyTech Finds / Setup Gamer",
        "creativeAngles": [
            "Sua placa de video esta torta? Evite peso no slot PCIe",
            "Setup gamer mais limpo com suporte ajustavel"
        ],
        "risk": "Medidas de gabinete variam; incluir guia de compatibilidade.",
    },
    {
        "id": "keycaps-pbt",
        "title": "Kit keycaps PBT para teclado mecanico",
        "niche": "Personalizacao de setup",
        "platform": "Mercado Livre / AliExpress",
        "sourceUrl": "https://www.mercadolivre.com.br/keycaps-para-teclados-mecnicos/p/MLB2046713672",
        "currentPrice": "Ofertas de R$28 a R$130+ conforme quantidade/perfil",
        "shipping": "Frete gratis em algumas ofertas; internacional exige atencao a prazo",
        "delivery": "Afiliado e mais simples ate validar demanda",
        "sellerReputation": "Algumas ofertas com +50 vendas; buscar kits com mais prova social antes de anuncio",
        "reviews": "Boa atratividade visual, mas nicho mais estetico",
        "returnPolicy": "Confirmar compatibilidade e devolucao na oferta escolhida",
        "operationModel": "Afiliado primeiro",
        "whySell": "Produto visual para criativos e publico gamer que gosta de customizacao.",
        "costTarget": "R$28 a R$70",
        "sellTarget": "R$79 a R$149",
        "margin": "Boa margem em dropshipping, mas exige conteudo claro de compatibilidade",
        "confidence": "Media",
        "wixCategory": "MobilyTech Finds / Setup Gamer",
        "creativeAngles": [
            "Teclado antigo, cara nova em minutos",
            "Personalize o setup sem comprar teclado novo"
        ],
        "risk": "Compatibilidade ABNT/ANSI/perfil gera devolucao se nao explicar bem.",
    },
    {
        "id": "teclado-mecanico-abnt2",
        "title": "Teclado gamer mecanico RGB ABNT2",
        "niche": "Perifericos gamer",
        "platform": "Mercado Livre / Amazon / Shopee",
        "sourceUrl": "https://www.mercadolivre.com.br/teclado-mecnico-gamer-rgb-abnt2-106-teclas-switch-red-silencioso-anti-ghosting-full-size-com-fio-usb-para-pc-notebook-portugus-brasil/p/MLB65123221",
        "currentPrice": "Preco varia bastante; validar SKU final antes de campanha",
        "shipping": "Produto maior; frete influencia conversao",
        "delivery": "Afiliado e mais seguro no inicio",
        "sellerReputation": "Busca indica bom custo-beneficio, mas tambem relatos de durabilidade a checar",
        "reviews": "Demanda alta por ABNT2/RGB; precisa filtrar qualidade",
        "returnPolicy": "Usar marketplace com politica clara de devolucao e garantia",
        "operationModel": "Afiliado primeiro",
        "whySell": "Periferico principal do setup; bom para combos com PCs e anuncios de retargeting.",
        "costTarget": "R$80 a R$180 conforme modelo",
        "sellTarget": "Afiliado: comissao; loja propria so apos fornecedor validado",
        "margin": "Melhor como afiliado por reduzir suporte/garantia",
        "confidence": "Media",
        "wixCategory": "MobilyTech Finds / Perifericos",
        "creativeAngles": [
            "Switch red, RGB e ABNT2 para jogar sem adaptacao",
            "O upgrade mais visivel do setup"
        ],
        "risk": "Garantia e suporte podem consumir margem se vendido como estoque proprio.",
    },
    {
        "id": "bias-light-led",
        "title": "Fita LED bias lighting USB para monitor",
        "niche": "Setup e conforto visual",
        "platform": "Mercado Livre / Shopee / AliExpress",
        "sourceUrl": "https://www.mercadolivre.com.br/5v-5050-tv-backlight-bias-lighting-usb-flexible-led-lig-1721/p/MLB2042722735",
        "currentPrice": "Oferta ML vista indisponivel; buscar fornecedor ativo antes de anunciar",
        "shipping": "Produto leve; bom para importacao se prazo for comunicado",
        "delivery": "Melhor como afiliado/teste ate achar fornecedor ativo",
        "sellerReputation": "Validacao fraca nesta rodada por oferta indisponivel",
        "reviews": "Boa ideia visual para criativo, mas precisa fornecedor com estoque",
        "returnPolicy": "Confirmar devolucao e funcionamento antes de publicar",
        "operationModel": "Testar afiliado primeiro",
        "whySell": "Produto barato e muito visual, bom para conteudo de setup.",
        "costTarget": "R$15 a R$35",
        "sellTarget": "R$49 a R$79",
        "margin": "Boa no papel, mas depende de fornecedor ativo e qualidade",
        "confidence": "Baixa-media",
        "wixCategory": "MobilyTech Finds / Setup Gamer",
        "creativeAngles": [
            "Deixe o monitor mais confortavel a noite",
            "Setup com cara gamer gastando pouco"
        ],
        "risk": "Fornecedor atual precisa ser trocado; nao subir anuncio antes de validar estoque.",
    },
]


AFFILIATE_LINKS = {
    "case-ssd-nvme": "https://meli.la/1DNJP2s",
    "case-ssd-sata": "https://meli.la/27TQAft",
    "fone-kz-castor": "https://meli.la/1WRhUoF",
    "kit-limpeza-esd": "https://meli.la/2qJ8vr9",
    "hub-usb-c": "https://meli.la/2GBaCfz",
    "suporte-gpu-antisag": "https://meli.la/1ycmCJ9",
    "keycaps-pbt": "https://meli.la/2tNDMiC",
    "teclado-mecanico-abnt2": "https://meli.la/2zS3ZBX",
}


MARKETPLACES = {
    "mercado-livre": {
        "name": "Mercado Livre",
        "logo": "assets/mercado-livre-logo.svg",
        "button": "Compre pelo Mercado Livre",
        "class": "market-ml",
    }
}


for item in FINALISTS:
    query = item["title"]
    affiliate_url = AFFILIATE_LINKS.get(item["id"])
    item["primaryMarketplace"] = "mercado-livre"
    item["marketplace"] = MARKETPLACES["mercado-livre"]
    item["affiliateUrl"] = affiliate_url or item["sourceUrl"]
    item["affiliateReady"] = bool(affiliate_url)
    item["affiliateButton"] = "Compre pelo Mercado Livre" if affiliate_url else "Ver oferta no Mercado Livre"
    item["affiliateStatus"] = (
        "Curadoria MobilyTech em marketplace confiavel."
        if affiliate_url
        else "Oferta selecionada para conferir no Mercado Livre."
    )
    item["whySell"] = SALES_COPY.get(item["id"], item["whySell"])
    item["publicPartnerNote"] = "Selecionado para complementar setups, upgrades e manutencao com compra segura."
    item["productImage"] = f"./assets/phase2-affiliates/{item['id']}.jpg"
    item["selectedCreativeVariant"] = 3 if affiliate_url else 2
    item["researchLinks"] = {
        "googleTrends": trends_link(query),
        "metaAdsLibrary": meta_link(query),
        "tikTokCreativeCenter": "https://ads.tiktok.com/business/creativecenter",
        "mercadoLivreTendencias": ml_trends_link(query),
        "mercadoLivreSearch": ml_search_link(query),
    }


def split_lines(text: str, limit: int = 34) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current: list[str] = []
    for word in words:
        if len(" ".join(current + [word])) > limit and current:
            lines.append(" ".join(current))
            current = [word]
        else:
            current.append(word)
    if current:
        lines.append(" ".join(current))
    return lines


def creative_svg(item: dict, variant: int) -> str:
    title_lines = split_lines(item["title"], 24)
    if variant == 1:
        headline = item["creativeAngles"][0]
        hook = "PROBLEMA"
        cta = "Validar oferta"
        color = "#16d9ff"
        glow = "#00ffc8"
    else:
        headline = item["creativeAngles"][1]
        hook = "BENEFICIO"
        cta = "Ver no site"
        color = "#7cf7ff"
        glow = "#3b82f6"
    headline_lines = split_lines(headline, 29)
    proof_lines = split_lines(item["sellerReputation"], 36)[:3]
    price_lines = split_lines(f"Preco visto: {item['currentPrice']}", 35)[:2]
    risk_lines = split_lines(item["risk"], 38)[:2]

    def tspans(lines: list[str], x: int, y: int, size: int, weight: int = 800, fill: str = "#f7fbff", gap: int = 36) -> str:
        parts = []
        for i, line in enumerate(lines):
            parts.append(f'<text x="{x}" y="{y + i * gap}" font-size="{size}" font-weight="{weight}" fill="{fill}">{html.escape(line)}</text>')
        return "\n".join(parts)

    icon = item["id"].split("-")[0].upper()
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1080" viewBox="0 0 1080 1080">
  <defs>
    <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0%" stop-color="#03070d"/>
      <stop offset="46%" stop-color="#071827"/>
      <stop offset="100%" stop-color="#020409"/>
    </linearGradient>
    <linearGradient id="panel" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0%" stop-color="{color}" stop-opacity="0.25"/>
      <stop offset="100%" stop-color="{glow}" stop-opacity="0.08"/>
    </linearGradient>
    <filter id="glow">
      <feGaussianBlur stdDeviation="12" result="blur"/>
      <feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge>
    </filter>
  </defs>
  <rect width="1080" height="1080" rx="54" fill="url(#bg)"/>
  <path d="M82 188 C250 78 422 125 535 196 C694 296 803 176 998 120" stroke="{color}" stroke-width="3" opacity="0.42" fill="none"/>
  <path d="M69 901 C244 792 397 905 536 846 C720 768 824 842 1014 748" stroke="{glow}" stroke-width="3" opacity="0.32" fill="none"/>
  <rect x="70" y="70" width="940" height="940" rx="42" fill="rgba(255,255,255,0.045)" stroke="{color}" stroke-opacity="0.52"/>
  <text x="98" y="130" font-size="30" font-weight="900" fill="{color}">MobilyTech BR</text>
  <text x="98" y="174" font-size="20" font-weight="800" fill="#b9f7ff">{html.escape(item["niche"])}</text>
  <rect x="774" y="96" width="190" height="48" rx="24" fill="{color}" fill-opacity="0.14" stroke="{color}" stroke-opacity="0.75"/>
  <text x="806" y="127" font-size="18" font-weight="900" fill="#f8ffff">{hook}</text>
  <circle cx="820" cy="395" r="178" fill="url(#panel)" stroke="{color}" stroke-width="4" filter="url(#glow)"/>
  <circle cx="820" cy="395" r="120" fill="#07111b" stroke="{glow}" stroke-opacity="0.55"/>
  <text x="820" y="410" font-size="58" font-weight="950" fill="{color}" text-anchor="middle">{icon}</text>
  <rect x="120" y="250" width="470" height="240" rx="32" fill="#07111b" fill-opacity="0.82" stroke="{color}" stroke-opacity="0.4"/>
  {tspans(headline_lines, 150, 313, 41, 950, "#ffffff", 48)}
  <rect x="120" y="530" width="835" height="180" rx="30" fill="#07111b" fill-opacity="0.74" stroke="#ffffff" stroke-opacity="0.08"/>
  <text x="150" y="585" font-size="23" font-weight="900" fill="{color}">PROVA / BENEFICIO</text>
  {tspans(proof_lines, 150, 628, 23, 800, "#e7f9ff", 32)}
  <rect x="120" y="748" width="520" height="138" rx="30" fill="{color}" fill-opacity="0.16" stroke="{color}" stroke-opacity="0.54"/>
  {tspans(price_lines, 150, 807, 24, 900, "#ffffff", 34)}
  <rect x="680" y="748" width="275" height="138" rx="30" fill="{glow}" fill-opacity="0.22" stroke="{glow}" stroke-opacity="0.72"/>
  <text x="818" y="829" font-size="28" font-weight="950" fill="#ffffff" text-anchor="middle">{cta}</text>
  <text x="120" y="932" font-size="18" font-weight="800" fill="#9fb7c7">{html.escape(" | ".join(risk_lines))}</text>
  <text x="120" y="970" font-size="16" font-weight="700" fill="#5edcff">Curadoria MobilyTech para setups, upgrades e manutencao.</text>
</svg>"""


def pil_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    font_name = "segoeuib.ttf" if bold else "segoeui.ttf"
    path = Path("C:/Windows/Fonts") / font_name
    if path.exists():
        return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def wrap_for_draw(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current: list[str] = []
    for word in words:
        candidate = " ".join(current + [word])
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width or not current:
            current.append(word)
        else:
            lines.append(" ".join(current))
            current = [word]
    if current:
        lines.append(" ".join(current))
    return lines


def draw_wrapped(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: str,
    max_width: int,
    line_gap: int,
    max_lines: int | None = None,
) -> int:
    lines = wrap_for_draw(draw, text, font, max_width)
    if max_lines:
        lines = lines[:max_lines]
    x, y = xy
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_gap
    return y


def creative_image(item: dict, variant: int, path: Path) -> None:
    W = H = 1080
    accent = "#17d9ff"
    green = "#00ffc6"
    bg = Image.new("RGB", (W, H), "#02050a")
    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    od.ellipse((-220, -180, 560, 520), fill=(23, 217, 255, 38))
    od.ellipse((610, 100, 1300, 790), fill=(0, 255, 198, 24))
    od.ellipse((220, 690, 1180, 1340), fill=(59, 130, 246, 20))
    overlay = overlay.filter(ImageFilter.GaussianBlur(34))
    bg = Image.alpha_composite(bg.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(bg)

    title_font = pil_font(48, True)
    mid_font = pil_font(32, True)
    body_font = pil_font(25, True)
    small_font = pil_font(21, True)
    tag_font = pil_font(24, True)

    if variant == 1:
        tag = "PROBLEMA DO SETUP"
        headline = item["creativeAngles"][0]
        sub = item["risk"]
        cta = "Veja a solucao"
        palette = accent
    elif variant == 2:
        tag = "DEMONSTRACAO"
        headline = item["creativeAngles"][1]
        sub = item["sellerReputation"]
        cta = "Curadoria MobilyTech"
        palette = green
    else:
        tag = "OFERTA VALIDADA" if item["affiliateReady"] else "VALIDAR OFERTA"
        headline = item["title"]
        sub = f"{item['currentPrice']} | {item['affiliateStatus']}"
        cta = item["affiliateButton"]
        palette = "#fff159"

    draw.rounded_rectangle((54, 54, 1026, 1026), radius=46, fill=(7, 18, 29, 228), outline=palette, width=3)
    for x in range(80, 1020, 72):
        draw.line((x, 70, x, 1010), fill=(121, 247, 255, 20), width=1)
    for y in range(80, 1020, 72):
        draw.line((70, y, 1010, y), fill=(121, 247, 255, 18), width=1)

    draw.rounded_rectangle((86, 88, 395, 144), radius=28, fill=(23, 217, 255, 32), outline=accent, width=2)
    draw.text((112, 103), "MobilyTech BR", font=tag_font, fill="#f8feff")
    draw.rounded_rectangle((720, 88, 982, 144), radius=28, fill=(255, 241, 89, 38), outline=palette, width=2)
    draw.text((750, 104), tag, font=small_font, fill="#f8feff")

    y = draw_wrapped(draw, (88, 188), headline, title_font, "#ffffff", 500, 54, 3)
    y = max(y + 18, 350)
    draw_wrapped(draw, (92, y), sub, body_font, "#dcecf5", 495, 32, 4)

    photo_path = ROOT / item["productImage"][2:]
    if photo_path.exists():
        product = Image.open(photo_path).convert("RGBA")
        product = ImageOps.contain(product, (340, 294), Image.Resampling.LANCZOS)
        card = Image.new("RGBA", (382, 336), (255, 255, 255, 0))
        cd = ImageDraw.Draw(card)
        cd.rounded_rectangle((0, 0, 382, 336), radius=34, fill=(247, 250, 252, 246), outline=(121, 247, 255, 120), width=2)
        card.alpha_composite(product, ((382 - product.width) // 2, (336 - product.height) // 2))
        bg.alpha_composite(card, (628, 226))

    draw.rounded_rectangle((92, 720, 988, 880), radius=34, fill=(5, 14, 23, 238), outline=palette, width=2)
    draw_wrapped(draw, (126, 756), item["currentPrice"], mid_font, "#ffffff", 805, 38, 2)
    draw_wrapped(draw, (126, 838), item["operationModel"], small_font, "#b9f7ff", 805, 28, 1)
    draw.rounded_rectangle((92, 914, 988, 984), radius=35, fill=palette, outline="#ffffff", width=1)
    draw.text((130, 936), cta[:42], font=mid_font, fill="#041018")

    path.parent.mkdir(parents=True, exist_ok=True)
    bg.convert("RGB").save(path, quality=92, optimize=True)


def write_creatives() -> None:
    CREATIVE_DIR.mkdir(parents=True, exist_ok=True)
    for item in FINALISTS:
        item["creatives"] = []
        for variant in (1, 2, 3):
            filename = f"{item['id']}-{variant:02d}.jpg"
            rel = f"./assets/phase2-creatives/{filename}"
            path = CREATIVE_DIR / filename
            creative_image(item, variant, path)
            selected = variant == item["selectedCreativeVariant"]
            item["creatives"].append(
                {
                    "variant": variant,
                    "file": rel,
                    "angle": (
                        item["creativeAngles"][variant - 1]
                        if variant <= len(item["creativeAngles"])
                        else f"Oferta e chamada para {item['marketplace']['name']}"
                    ),
                    "status": "selecionado para vitrine" if selected else "alternativa de criativo",
                    "selected": selected,
                }
            )
        item["selectedCreative"] = next(creative["file"] for creative in item["creatives"] if creative["selected"])


def brand_logo_svg(brand: dict) -> str:
    name = html.escape(brand["name"])
    accent = brand["accent"]
    size = 38 if len(brand["name"]) <= 5 else 30 if len(brand["name"]) <= 8 else 25
    mark = ""
    if brand["id"] == "microsoft":
        mark = """
  <rect x="34" y="29" width="14" height="14" fill="#f25022"/>
  <rect x="51" y="29" width="14" height="14" fill="#7fba00"/>
  <rect x="34" y="46" width="14" height="14" fill="#00a4ef"/>
  <rect x="51" y="46" width="14" height="14" fill="#ffb900"/>"""
    elif brand["id"] == "amd":
        mark = """
  <path d="M35 28h31v11H46v20H35z" fill="#22c55e"/>
  <path d="M67 28h18v18H74v-7h-7z" fill="#22c55e"/>"""
    elif brand["id"] == "nvidia":
        mark = """
  <path d="M34 45c13-17 34-17 47 0-13 17-34 17-47 0z" fill="none" stroke="#76b900" stroke-width="5"/>
  <circle cx="58" cy="45" r="8" fill="#76b900"/>"""
    elif brand["id"] == "kingston":
        mark = """
  <path d="M55 23c14 9 18 24 11 41H39c-7-17-3-32 16-41z" fill="#ef4444" opacity=".9"/>
  <path d="M45 55c8 5 17 5 26 0" stroke="#fff" stroke-width="4" stroke-linecap="round"/>"""
    else:
        mark = f"""
  <circle cx="55" cy="45" r="24" fill="none" stroke="{accent}" stroke-width="4" opacity=".9"/>
  <path d="M42 45h26" stroke="{accent}" stroke-width="4" stroke-linecap="round"/>"""
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="240" height="88" viewBox="0 0 240 88" role="img" aria-label="{name}">
  <rect width="240" height="88" rx="14" fill="#121820"/>
  <rect x="1" y="1" width="238" height="86" rx="13" fill="none" stroke="{accent}" stroke-opacity=".5"/>
  {mark}
  <text x="134" y="54" text-anchor="middle" font-family="Arial, Helvetica, sans-serif" font-size="{size}" font-weight="900" fill="#f8fbff" letter-spacing="0">{name}</text>
</svg>"""


def write_brand_logos() -> None:
    BRAND_OFFICIAL_DIR.mkdir(parents=True, exist_ok=True)
    mobily_logo = ROOT / "assets" / "mobilytech-logo.png"
    mobily_official = BRAND_OFFICIAL_DIR / "mobilytech.png"
    if mobily_logo.exists() and not mobily_official.exists():
        mobily_official.write_bytes(mobily_logo.read_bytes())
    for brand in BRANDS:
        official = BRAND_OFFICIAL_DIR / f"{brand['id']}.svg"
        fallback = BRAND_LOGO_DIR / f"{brand['id']}.svg"
        if official.exists():
            continue
        if fallback.exists():
            official.write_text(fallback.read_text(encoding="utf-8"), encoding="utf-8")
        else:
            official.write_text(brand_logo_svg(brand), encoding="utf-8")


def brand_logo_cards(prefix: str = "./") -> str:
    return "\n".join(
        f'          <span class="brand-pill" style="--brand-accent:{brand["accent"]}"><img src="{prefix}assets/brand-officials/{brand["id"]}.svg" alt="{html.escape(brand["name"])}"></span>'
        for brand in BRANDS
    )


def write_phase2_json() -> Path:
    payload = {
        "generatedAt": GENERATED_AT,
        "status": "draft-for-approval-no-paid-ads",
        "strategy": "Vercel storefront with Wix/Premium bridge planning.",
        "backupPath": str(BACKUP_PATH),
        "wixPremiumSiteId": WIX_PREMIUM_SITE_ID,
        "wixDomain": WIX_DOMAIN,
        "vercelProduction": VERCEL_PRODUCTION,
        "sources": [
            "https://tendencias.mercadolivre.com.br/",
            "https://www.facebook.com/ads/library",
            "https://ads.tiktok.com/business/creativecenter",
            "https://ads.tiktok.com/help/article/top-products?lang=en",
            "https://www.gov.br/mj/pt-br/assuntos/noticias/consumidor-tem-direito-ao-arrependimento-em-compras-on-line",
            "https://www.planalto.gov.br/ccivil_03/leis/l8078compilado.htm",
        ],
        "finalists": FINALISTS,
    }
    path = DATA_DIR / "phase2-finalists.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_page() -> Path:
    html_doc = r"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>MobilyTech BR | PCs revisados e hardware</title>
  <meta name="description" content="MobilyTech BR: PCs revisados, hardware, montagem sob orcamento, limpeza de PCs e curadoria MobilyTech Finds." />
  <link rel="icon" href="./assets/favicon.png" />
  <style>
    :root {
      color-scheme: dark;
      --bg: #02050a;
      --bg-2: #07111c;
      --panel: rgba(8, 23, 34, 0.78);
      --panel-2: rgba(10, 35, 51, 0.86);
      --line: rgba(118, 242, 255, 0.22);
      --cyan: #17d9ff;
      --cyan-2: #79f7ff;
      --blue: #2e7dff;
      --green: #00ffc6;
      --text: #f6fbff;
      --muted: #a9bed0;
      --radius: 22px;
      font-family: Nunito, Inter, Segoe UI, Arial, sans-serif;
    }
    * { box-sizing: border-box; }
    html { scroll-behavior: smooth; }
    body {
      margin: 0;
      min-height: 100vh;
      background:
        radial-gradient(circle at 15% 10%, rgba(23, 217, 255, 0.13), transparent 28rem),
        radial-gradient(circle at 82% 18%, rgba(0, 255, 198, 0.09), transparent 26rem),
        linear-gradient(180deg, #03070d 0%, #06101b 44%, #02050a 100%);
      color: var(--text);
      letter-spacing: 0;
      overflow-x: hidden;
    }
    a { color: inherit; text-decoration: none; }
    button, input { font: inherit; }
    .shell { width: min(1510px, calc(100% - 36px)); margin: 0 auto; }
    .promo-strip {
      background: rgba(4, 10, 16, 0.86);
      border-bottom: 1px solid var(--line);
      color: #d8fbff;
      font-size: 13px;
      font-weight: 800;
    }
    .promo-strip .shell {
      min-height: 38px;
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 18px;
      text-align: center;
    }
    .promo-strip strong { color: var(--green); }
    .top-nav {
      position: sticky;
      top: 0;
      z-index: 10;
      background: rgba(2, 7, 13, 0.88);
      backdrop-filter: blur(18px);
      border-bottom: 1px solid var(--line);
    }
    .nav-row {
      height: 78px;
      display: grid;
      grid-template-columns: 180px minmax(0, 1fr) 220px 90px;
      align-items: center;
      gap: 8px;
    }
    .brand { display: flex; align-items: center; gap: 10px; min-width: 0; }
    .brand img { width: 40px; height: 40px; object-fit: contain; filter: drop-shadow(0 0 10px rgba(23,217,255,.42)); }
    .brand span { font-size: 18px; font-weight: 950; white-space: nowrap; }
    .menu {
      display: flex;
      align-items: center;
      justify-content: flex-start;
      gap: 7px;
      overflow-x: auto;
      scrollbar-width: none;
    }
    .menu::-webkit-scrollbar { display: none; }
    .menu a {
      min-height: 36px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 5px;
      font-size: 11.5px;
      font-weight: 900;
      color: #e4f8ff;
      white-space: nowrap;
      opacity: .95;
      padding: 0 8px;
      border: 1px solid rgba(121,247,255,.16);
      border-radius: 999px;
      background: rgba(255,255,255,.045);
    }
    .menu a:hover { color: var(--cyan); }
    .nav-icon {
      width: 14px;
      height: 14px;
      display: inline-grid;
      place-items: center;
      color: var(--cyan);
      flex: 0 0 auto;
    }
    .nav-icon svg {
      width: 14px;
      height: 14px;
      fill: none;
      stroke: currentColor;
      stroke-width: 1.9;
      stroke-linecap: round;
      stroke-linejoin: round;
    }
    .search {
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 0 14px;
      height: 42px;
      border-radius: 999px;
      background: rgba(255,255,255,.07);
      border: 1px solid rgba(255,255,255,.12);
    }
    .search input {
      width: 100%;
      background: transparent;
      border: 0;
      outline: 0;
      color: var(--text);
      font-weight: 800;
      min-width: 0;
    }
    .cart-btn, .primary, .ghost {
      display: inline-grid;
      place-items: center;
      border: 0;
      cursor: pointer;
      min-height: 42px;
      border-radius: 999px;
      font-weight: 950;
      color: #041018;
      background: linear-gradient(135deg, var(--cyan), var(--green));
      box-shadow: 0 0 26px rgba(23,217,255,.22);
      padding: 0 18px;
      line-height: 1;
      text-align: center;
      text-decoration: none;
      white-space: nowrap;
      vertical-align: middle;
    }
    .ghost {
      background: transparent;
      color: var(--cyan-2);
      border: 1px solid rgba(121,247,255,.42);
      box-shadow: none;
    }
    .hero {
      padding: 30px 0 18px;
    }
    .hero-panel {
      min-height: 430px;
      border-radius: 28px;
      overflow: hidden;
      position: relative;
      background:
        linear-gradient(105deg, rgba(3,7,13,.96) 0%, rgba(4,26,39,.92) 43%, rgba(3,7,13,.68) 100%),
        url("./assets/cleaning-neon-bg.png") center/cover;
      border: 1px solid rgba(121,247,255,.18);
      box-shadow: 0 22px 65px rgba(0,0,0,.28);
    }
    .hero-panel::after {
      content: "";
      position: absolute;
      inset: 0;
      background-image:
        linear-gradient(rgba(121,247,255,.08) 1px, transparent 1px),
        linear-gradient(90deg, rgba(121,247,255,.07) 1px, transparent 1px);
      background-size: 46px 46px;
      mask-image: linear-gradient(90deg, #000 0%, transparent 68%);
      pointer-events: none;
    }
    .hero-content {
      position: relative;
      z-index: 1;
      display: grid;
      grid-template-columns: minmax(0, 1fr) 600px;
      gap: 24px;
      min-height: 430px;
      align-items: center;
      padding: 54px 64px;
    }
    .eyebrow {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 7px 12px;
      border-radius: 999px;
      background: rgba(23,217,255,.12);
      border: 1px solid rgba(23,217,255,.3);
      color: var(--cyan-2);
      font-size: 12px;
      font-weight: 950;
      text-transform: uppercase;
    }
    h1 {
      margin: 18px 0 12px;
      font-size: clamp(44px, 5vw, 78px);
      line-height: .94;
      letter-spacing: 0;
      max-width: 760px;
    }
    .lead {
      max-width: 650px;
      color: #d8e8f4;
      font-size: 18px;
      line-height: 1.5;
      font-weight: 750;
    }
    .hero-actions { display: flex; flex-wrap: wrap; gap: 12px; margin-top: 26px; }
    .hero-actions .primary, .hero-actions .ghost { min-width: 178px; min-height: 54px; font-size: 16px; }
    .hero-stage {
      min-height: 340px;
      position: relative;
      display: grid;
      place-items: center;
    }
    .hero-stage img {
      position: absolute;
      max-width: 92%;
      max-height: 330px;
      object-fit: contain;
      filter:
        drop-shadow(0 16px 0 rgba(255,255,255,.9))
        drop-shadow(0 24px 28px rgba(0,255,198,.12))
        drop-shadow(0 0 32px rgba(23,217,255,.28));
      transition: opacity .35s ease, transform .35s ease;
    }
    .hero-card-mini {
      position: absolute;
      right: 10px;
      bottom: 8px;
      width: 260px;
      border-radius: 18px;
      background: rgba(3, 12, 19, .76);
      border: 1px solid rgba(121,247,255,.22);
      padding: 16px;
      box-shadow: 0 12px 34px rgba(0,0,0,.3);
    }
    .hero-card-mini strong { color: var(--green); font-size: 22px; display: block; }
    .section { padding: 34px 0; }
    .section-head {
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 18px;
    }
    .section h2 {
      margin: 0;
      font-size: clamp(28px, 3vw, 46px);
      line-height: 1;
    }
    .section p.sub {
      margin: 8px 0 0;
      color: var(--muted);
      font-weight: 800;
      max-width: 740px;
    }
    .deal-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 18px;
    }
    .deal-card, .feature-card, .review-card, .finalist-card, .creative-card, .policy-card {
      border-radius: var(--radius);
      background: var(--panel);
      border: 1px solid rgba(121,247,255,.18);
      box-shadow: 0 18px 48px rgba(0,0,0,.24);
      overflow: hidden;
      min-width: 0;
    }
    .deal-card {
      padding: 16px;
      display: flex;
      flex-direction: column;
      min-height: 330px;
    }
    .deal-img {
      height: 166px;
      display: grid;
      place-items: center;
      border-radius: 18px;
      background: radial-gradient(circle at 52% 40%, rgba(23,217,255,.22), rgba(0,0,0,.08) 54%, rgba(0,0,0,.28));
      overflow: hidden;
    }
    .deal-img img { max-width: 96%; max-height: 154px; object-fit: contain; filter: drop-shadow(0 16px 20px rgba(0,0,0,.32)); }
    .deal-card h3 { margin: 14px 0 8px; font-size: 17px; line-height: 1.18; min-height: 40px; }
    .price { color: var(--cyan); font-size: 25px; font-weight: 950; margin-top: auto; }
    .specs { color: var(--muted); font-size: 12px; font-weight: 800; line-height: 1.45; margin: 0 0 12px; }
    .deal-actions { display: flex; gap: 8px; margin-top: 12px; }
    .deal-actions button, .deal-actions a { flex: 1; display: grid; place-items: center; min-height: 38px; border-radius: 999px; font-size: 12px; font-weight: 950; }
    .feature-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
    }
    .feature-card {
      min-height: 300px;
      display: grid;
      grid-template-columns: 1fr 310px;
      align-items: center;
      padding: 34px;
      position: relative;
      background:
        linear-gradient(135deg, rgba(8,33,46,.92), rgba(3,7,13,.8)),
        url("./assets/brazil-purchase-card-bg.png") center/cover;
    }
    .feature-card:nth-child(2) { background-image: linear-gradient(135deg, rgba(4, 64, 55, .9), rgba(3,7,13,.72)), url("./assets/cleaning-neon-bg.png"); }
    .feature-card h3 { font-size: 34px; line-height: 1.02; margin: 0 0 12px; }
    .feature-card p { color: #def6ff; font-weight: 800; line-height: 1.42; margin: 0 0 18px; }
    .feature-card img { max-width: 310px; max-height: 260px; object-fit: contain; justify-self: end; filter: drop-shadow(0 22px 22px rgba(0,0,0,.34)); }
    .finalist-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 18px;
    }
    .finalist-card { padding: 14px; display: flex; flex-direction: column; gap: 10px; min-height: 388px; }
    .affiliate-photo {
      width: 100%;
      aspect-ratio: 1 / .78;
      display: grid;
      place-items: center;
      border-radius: 18px;
      overflow: hidden;
      background: linear-gradient(145deg, rgba(255,255,255,.98), rgba(219,241,247,.9));
      border: 1px solid rgba(121,247,255,.2);
      box-shadow: inset 0 0 0 1px rgba(255,255,255,.5), 0 16px 28px rgba(0,0,0,.18);
    }
    .affiliate-photo img { width: auto; height: auto; max-width: 92%; max-height: 88%; object-fit: contain; display: block; border-radius: 12px; }
    .market-row { display: flex; align-items: center; justify-content: space-between; gap: 8px; }
    .market-badge {
      display: inline-flex;
      align-items: center;
      gap: 7px;
      min-width: 0;
      border-radius: 999px;
      padding: 5px 9px;
      color: #071018;
      background: linear-gradient(135deg, #ffe600, #fff2a8);
      font-size: 10.5px;
      font-weight: 950;
      box-shadow: 0 0 18px rgba(255,230,0,.15);
    }
    .market-badge img { width: 18px; height: 18px; object-fit: contain; }
    .affiliate-status { font-size: 10px; color: #9fb3bd; font-weight: 850; line-height: 1.25; min-height: 26px; }
    .market-button {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      width: 100%;
      min-height: 42px;
      margin-top: auto;
      border-radius: 999px;
      border: 1px solid rgba(255,255,255,.2);
      color: #071018;
      font-size: 12px;
      font-weight: 950;
      text-decoration: none;
      text-align: center;
      padding: 0 12px;
      box-shadow: 0 0 26px rgba(255,230,0,.16), inset 0 1px 0 rgba(255,255,255,.45);
      backdrop-filter: blur(12px);
    }
    .market-button img { width: 27px; height: 27px; object-fit: contain; flex: 0 0 auto; }
    .market-button.market-ml { background: linear-gradient(135deg, #fff159 0%, #ffe000 42%, #28a8ff 120%); }
    .badge {
      align-self: flex-start;
      border-radius: 999px;
      background: rgba(23,217,255,.12);
      border: 1px solid rgba(23,217,255,.32);
      color: var(--cyan-2);
      font-size: 11px;
      font-weight: 950;
      padding: 5px 9px;
    }
    .finalist-card h3 { font-size: 15px; line-height: 1.18; margin: 0; }
    .finalist-card p { margin: 0; color: var(--muted); font-size: 12px; font-weight: 800; line-height: 1.35; }
    .finalist-card .model { color: var(--green); }
    .creative-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 16px;
    }
    .creative-card { padding: 12px; background: rgba(6,17,27,.86); }
    .creative-card img { width: 100%; aspect-ratio: 1/1; object-fit: cover; border-radius: 16px; display: block; background: #03101a; }
    .creative-card span { display: block; margin-top: 9px; font-size: 12px; font-weight: 900; color: #dffbff; }
    .reviews-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 18px;
    }
    .review-card { padding: 22px; min-height: 178px; background: rgba(255,255,255,.94); color: #09101a; }
    .review-card .stars { color: #ffb300; font-size: 20px; letter-spacing: 0; }
    .review-card h3 { margin: 8px 0; font-size: 18px; }
    .review-card p { margin: 0; color: #243244; font-weight: 750; line-height: 1.4; }
    .brand-wall {
      border-radius: 30px;
      padding: clamp(22px, 4vw, 42px);
      background:
        radial-gradient(circle at 20% 0%, rgba(23,217,255,.14), transparent 34%),
        linear-gradient(145deg, rgba(10,18,28,.94), rgba(3,8,14,.96));
      border: 1px solid rgba(121,247,255,.2);
      box-shadow: 0 22px 58px rgba(0,0,0,.32), inset 0 0 0 1px rgba(255,255,255,.035);
    }
    .brand-wall h2 { text-align: center; margin-bottom: 26px; }
    .brand-row {
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: nowrap;
      gap: clamp(16px, 2vw, 30px);
      overflow-x: auto;
      scrollbar-width: none;
    }
    .brand-row::-webkit-scrollbar { display: none; }
    .brand-pill {
      flex: 0 0 auto;
      min-width: 74px;
      display: grid;
      place-items: center;
      padding: 0;
      border: 0;
      border-radius: 0;
      background: transparent;
      box-shadow: none;
      overflow: visible;
    }
    .brand-pill:nth-last-child(2), .brand-pill:last-child { grid-column: auto; }
    .brand-pill img { display: block; width: auto; height: 32px; max-width: 112px; max-height: 32px; object-fit: contain; filter: drop-shadow(0 0 10px rgba(23,217,255,.22)); }
    .policy-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 16px;
    }
    .policy-card { padding: 20px; }
    .policy-card h3 { margin: 0 0 10px; color: var(--cyan); }
    .policy-card p { margin: 0; color: #d9e8f3; line-height: 1.45; font-weight: 760; }
    footer {
      margin-top: 40px;
      padding: 44px 0;
      background: #03070d;
      border-top: 1px solid rgba(121,247,255,.16);
    }
    .footer-grid {
      display: grid;
      grid-template-columns: 1.5fr repeat(4, 1fr);
      gap: 26px;
    }
    footer h3 { margin: 0 0 12px; color: #fff; font-size: 16px; }
    footer a, footer p { display: block; color: #c3d4df; font-weight: 800; font-size: 14px; margin: 8px 0; }
    .drawer {
      position: fixed;
      inset: 0 0 0 auto;
      z-index: 30;
      width: min(430px, 100%);
      background: rgba(4,11,18,.96);
      border-left: 1px solid rgba(121,247,255,.22);
      transform: translateX(110%);
      transition: transform .25s ease;
      padding: 22px;
      overflow-y: auto;
      box-shadow: -22px 0 60px rgba(0,0,0,.38);
    }
    .drawer.open { transform: translateX(0); }
    .drawer h2 { margin-top: 0; }
    .drawer-head {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: flex-start;
      padding-bottom: 14px;
      border-bottom: 1px solid rgba(121,247,255,.16);
    }
    .drawer-head h2 { margin: 0; font-size: 28px; line-height: 1; }
    .drawer-kicker {
      display: block;
      margin-bottom: 6px;
      color: var(--green);
      font-size: 11px;
      font-weight: 950;
      text-transform: uppercase;
    }
    .drawer-close { min-height: 36px; padding: 0 14px; font-size: 12px; }
    .option {
      display: flex;
      gap: 10px;
      align-items: flex-start;
      padding: 11px;
      border-radius: 14px;
      background: rgba(255,255,255,.055);
      margin: 8px 0;
      font-weight: 850;
      color: #dff7ff;
    }
    .cart-list { display: grid; gap: 12px; margin: 18px 0; }
    .cart-line {
      display: grid;
      grid-template-columns: 72px 1fr;
      gap: 12px;
      border: 1px solid rgba(121,247,255,.15);
      border-radius: 18px;
      padding: 12px;
      background: rgba(255,255,255,.055);
      color: #dff7ff;
      font-weight: 800;
    }
    .cart-line img {
      width: 72px;
      height: 72px;
      object-fit: contain;
      border-radius: 14px;
      background: radial-gradient(circle, rgba(23,217,255,.15), rgba(0,0,0,.28));
    }
    .cart-line strong { display: block; line-height: 1.18; }
    .cart-options { margin: 6px 0 8px; color: var(--muted); font-size: 11px; line-height: 1.35; }
    .cart-line-bottom { display: flex; align-items: center; justify-content: space-between; gap: 10px; }
    .cart-remove {
      border: 0;
      background: transparent;
      color: #79f7ff;
      cursor: pointer;
      font-weight: 950;
      padding: 0;
    }
    .checkout-card {
      display: grid;
      gap: 12px;
      margin-top: 14px;
      padding: 16px;
      border-radius: 20px;
      border: 1px solid rgba(121,247,255,.18);
      background: rgba(255,255,255,.06);
    }
    .checkout-row { display: flex; align-items: center; justify-content: space-between; gap: 12px; color: #dff7ff; font-weight: 900; }
    .checkout-row strong { color: var(--cyan); font-size: 22px; }
    .payment-chips { display: flex; flex-wrap: wrap; gap: 8px; }
    .payment-chips span {
      display: inline-grid;
      place-items: center;
      min-height: 28px;
      border-radius: 999px;
      padding: 0 10px;
      border: 1px solid rgba(121,247,255,.2);
      color: #dffbff;
      background: rgba(23,217,255,.08);
      font-size: 11px;
      font-weight: 950;
    }
    .cart-actions { display: grid; gap: 10px; }
    .cart-actions .primary, .cart-actions .ghost { width: 100%; min-height: 44px; }
    .mini-field {
      width: 100%;
      min-height: 42px;
      border-radius: 999px;
      border: 1px solid rgba(121,247,255,.22);
      background: rgba(0,0,0,.22);
      color: var(--text);
      padding: 0 14px;
      outline: 0;
      font-weight: 850;
    }
    .sr-note {
      margin-top: 12px;
      color: #8fafbf;
      font-size: 12px;
      font-weight: 800;
      line-height: 1.4;
    }
    @media (max-width: 1120px) {
      .nav-row { grid-template-columns: 170px 1fr 230px 104px; gap: 12px; }
      .hero-content { grid-template-columns: 1fr 430px; padding: 42px; }
      .deal-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .finalist-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .brand-row { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    }
    @media (max-width: 760px) {
      .shell { width: min(100% - 22px, 720px); }
      .promo-strip .shell {
        min-height: 30px;
        font-size: 10px;
        line-height: 1.15;
        gap: 8px;
      }
      .nav-row {
        height: auto;
        grid-template-columns: 1fr auto;
        grid-template-areas:
          "brand cart"
          "search search"
          "menu menu";
        padding: 10px 0;
      }
      .brand { grid-area: brand; }
      .brand span { font-size: 16px; }
      .brand img { width: 34px; height: 34px; }
      .menu {
        grid-area: menu;
        justify-content: start;
        gap: 10px;
        padding: 1px 18px 1px 0;
        mask-image: linear-gradient(90deg, #000 0%, #000 calc(100% - 24px), transparent 100%);
      }
      .menu a { min-height: 31px; font-size: 10.5px; padding: 0 10px; }
      .nav-icon, .nav-icon svg { width: 14px; height: 14px; }
      .search { grid-area: search; height: 36px; }
      .search input { font-size: 14px; }
      .cart-btn { grid-area: cart; min-height: 38px; padding: 0 13px; font-size: 12px; }
      .hero { padding-top: 14px; }
      .hero-panel { min-height: auto; border-radius: 20px; }
      .hero-content { grid-template-columns: 1fr; padding: 24px 18px 18px; gap: 14px; min-height: 0; }
      .eyebrow { font-size: 10px; padding: 6px 10px; }
      h1 { font-size: 28px; line-height: 1.02; margin: 14px 0 10px; }
      .lead { font-size: 12.5px; line-height: 1.38; max-width: 100%; }
      .hero-actions { gap: 8px; margin-top: 18px; }
      .hero-stage {
        min-height: 130px;
        display: grid;
        grid-template-columns: 40% minmax(0, 1fr);
        gap: 10px;
        align-items: center;
      }
      .hero-stage img {
        position: relative;
        max-width: 100%;
        max-height: 128px;
        justify-self: center;
      }
      .hero-card-mini {
        position: relative;
        right: auto;
        bottom: auto;
        width: 100%;
        margin-top: 0;
        padding: 12px;
        border-radius: 16px;
      }
      .hero-card-mini span { font-size: 13px; }
      .hero-card-mini strong { font-size: 18px; line-height: 1.05; }
      .hero-card-mini .specs { font-size: 10.5px; margin-bottom: 0; }
      .hero-actions .primary, .hero-actions .ghost { min-width: 0; min-height: 36px; font-size: 11px; padding: 0 12px; }
      .section { padding: 24px 0; }
      .section h2 { font-size: 25px; line-height: 1.05; }
      .section p.sub { font-size: 12px; line-height: 1.35; }
      .section-head { align-items: start; flex-direction: column; }
      .deal-grid, .finalist-grid, .creative-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 10px;
      }
      .deal-card { min-height: 260px; padding: 10px; border-radius: 16px; }
      .deal-img { height: 112px; border-radius: 14px; }
      .deal-img img { max-height: 104px; }
      .deal-card h3 { font-size: 12px; min-height: 34px; }
      .price { font-size: 18px; }
      .specs, .finalist-card p { font-size: 10.5px; }
      .deal-actions { flex-direction: column; gap: 6px; }
      .deal-actions button, .deal-actions a { min-height: 32px; font-size: 10.5px; }
      .feature-grid { grid-template-columns: 1fr; }
      .feature-card { grid-template-columns: 1fr 124px; min-height: 200px; padding: 18px; border-radius: 18px; }
      .feature-card h3 { font-size: 22px; }
      .feature-card p { font-size: 12px; }
      .feature-card img { max-width: 130px; max-height: 140px; }
      .finalist-card { min-height: 318px; padding: 10px; border-radius: 16px; }
      .affiliate-photo { border-radius: 13px; }
      .market-row { align-items: flex-start; flex-direction: column; gap: 6px; }
      .market-button { min-height: 34px; font-size: 10px; gap: 5px; white-space: normal; line-height: 1.12; padding: 5px 8px; }
      .market-button img { width: 21px; height: 21px; }
      .affiliate-status { font-size: 9px; min-height: 0; }
      .finalist-card h3 { font-size: 12px; }
      .finalist-card p,
      .creative-card span {
        overflow-wrap: anywhere;
        display: -webkit-box;
        -webkit-box-orient: vertical;
        overflow: hidden;
      }
      .finalist-card p { -webkit-line-clamp: 3; }
      .finalist-card .model { -webkit-line-clamp: 2; }
      .creative-card span { -webkit-line-clamp: 2; font-size: 10.5px; line-height: 1.22; }
      .badge { font-size: 9px; padding: 4px 7px; }
      .reviews-grid, .policy-grid { grid-template-columns: 1fr; }
      .brand-wall { padding: 20px 12px; }
      .brand-row { display: flex; flex-wrap: nowrap; gap: 18px; overflow-x: auto; }
      .brand-pill, .brand-pill:nth-last-child(2), .brand-pill:last-child { grid-column: auto; }
      .brand-pill { min-width: 78px; }
      .brand-pill img { height: 27px; max-width: 92px; max-height: 27px; }
      .footer-grid { grid-template-columns: 1fr 1fr; gap: 18px; }
      footer .brand-block { grid-column: 1 / -1; }
    }
  </style>
</head>
<body>
  <div class="promo-strip">
    <div class="shell">
      <span>MobilyTech BR: <strong>PCs revisados + MobilyTech Finds</strong></span>
      <span>Ofertas e campanhas com curadoria</span>
    </div>
  </div>
  <nav class="top-nav">
    <div class="shell nav-row">
      <a class="brand" href="#inicio" aria-label="MobilyTech BR">
        <img src="./assets/mobilytech-logo.png" alt="" />
        <span>MobilyTech BR</span>
      </a>
      <div class="menu" aria-label="Navegacao principal">
        <a href="./fase2/ofertas.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M4 7h16v4H4z"/><path d="M6 11h12v8H6z"/><path d="M9 7V5h6v2"/></svg></span>Ofertas</a>
        <a href="./fase2/ofertas.html#pcs"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M5 8h14v8H5z"/><path d="M9 20h6"/><path d="M12 16v4"/><path d="M8 11h2m4 0h2"/></svg></span>PC Gamer</a>
        <a href="./fase2/montagem.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M5 8h14v8H5z"/><path d="M9 20h6"/><path d="M12 16v4"/><path d="M8 11h2m4 0h2"/></svg></span>Monte seu PC</a>
        <a href="./fase2/limpeza.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M6 14c3-1 4-4 4-9 3 2 5 5 6 9"/><path d="M5 14h14l-1 6H6z"/><path d="M8 17h8"/></svg></span>Limpeza</a>
        <a href="./fase2/achados.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="m12 3 1.8 5.4 5.2 1.8-5.2 1.8L12 17.5 10.2 12 5 10.2l5.2-1.8z"/><path d="m18 15 .7 2.1 2.1.7-2.1.7-.7 2.1-.7-2.1-2.1-.7 2.1-.7z"/></svg></span>MobilyTech Finds</a>
        <a href="./fase2/avaliacoes.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M12 3 9.2 8.8 3 9.7l4.5 4.4-1.1 6.2L12 17.4l5.6 2.9-1.1-6.2L21 9.7l-6.2-.9z"/></svg></span>Avaliacoes</a>
        <a href="./fase2/contato.html"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><path d="M4 6h16v12H4z"/><path d="m4 7 8 6 8-6"/></svg></span>Suporte</a>
      </div>
      <label class="search" aria-label="Pesquisar">
        <span>⌕</span>
        <input id="searchInput" placeholder="Buscar PC, SSD, limpeza..." />
      </label>
      <button class="cart-btn" id="cartButton">Carrinho <span id="cartCount">0</span></button>
    </div>
  </nav>

  <main id="inicio">
    <section class="hero shell">
      <div class="hero-panel">
        <div class="hero-content">
          <div>
            <span class="eyebrow">Confianca que voce sente</span>
            <h1>PCs revisados, upgrades e achados tech com garantia.</h1>
            <p class="lead">A MobilyTech BR une PCs reais revisados, hardware selecionado, limpeza especializada, montagem sob orcamento e curadoria de produtos para setup.</p>
            <div class="hero-actions">
              <a class="primary" href="./fase2/ofertas.html">Ver PCs e ofertas</a>
              <a class="ghost" href="./fase2/achados.html">Ver MobilyTech Finds</a>
            </div>
          </div>
          <div class="hero-stage">
            <img id="heroProduct" src="./assets/generated/pcryzen-5-3600-cutout.png" alt="PC Gamer MobilyTech" />
            <div class="hero-card-mini">
              <span>Julho Tech</span>
              <strong>PCs + upgrades</strong>
              <p class="specs">Ofertas reais, upgrades claros e atendimento humano.</p>
            </div>
          </div>
        </div>
      </div>
    </section>

    <section id="ofertas" class="section shell">
      <div class="section-head">
        <div>
          <h2>Promocoes atuais MobilyTech</h2>
          <p class="sub">PCs e hardware reais do catalogo atual, com configuracao, carrinho e atendimento pelos canais da MobilyTech.</p>
        </div>
        <a class="ghost" href="./fase2/ofertas.html">Abrir catalogo completo</a>
      </div>
      <div id="productGrid" class="deal-grid"></div>
    </section>

    <section id="monte" class="section shell">
      <div class="feature-grid">
        <article class="feature-card">
          <div>
            <span class="eyebrow">Custom gaming PCs</span>
            <h3>Monte seu PC sob orcamento</h3>
            <p>Escolha uso, jogos, limite de preco e upgrades. A montagem continua personalizada, com garantia informada antes da compra.</p>
            <a class="primary" href="./fase2/montagem.html">Consultar orcamento</a>
          </div>
          <img src="./assets/assembly-pc-build-cutout.png" alt="Montagem de PC" />
        </article>
        <article id="limpeza" class="feature-card">
          <div>
            <span class="eyebrow">PC cleaning service</span>
            <h3>Limpeza e relatorio</h3>
            <p>Servico de limpeza com foco em cuidado, organizacao e relatorio visual do antes/depois, mantendo a pegada de pos-venda.</p>
            <a class="primary" href="./fase2/limpeza.html">Agendar limpeza</a>
          </div>
          <img src="./assets/pc-cleaning-service-cutout.png" alt="Limpeza de PC" />
        </article>
      </div>
    </section>

    <section id="achados" class="section shell">
      <div class="section-head">
        <div>
          <h2>MobilyTech Finds selecionados</h2>
          <p class="sub">Produtos tech escolhidos para completar setups, upgrades e manutencao com compra segura em marketplaces.</p>
        </div>
        <a class="ghost" href="./fase2/achados.html">Abrir MobilyTech Finds</a>
      </div>
      <div id="finalistGrid" class="finalist-grid"></div>
    </section>

    <section id="avaliacoes" class="section shell">
      <div class="section-head">
        <div>
          <h2>Avaliacoes da MobilyTech BR</h2>
          <p class="sub">Blocos no estilo iBUYPOWER, usando os canais reais de confianca da MobilyTech.</p>
        </div>
      </div>
      <div class="reviews-grid">
        <article class="review-card">
          <div class="stars">★★★★★</div>
          <h3>OLX</h3>
          <p>Avaliacoes publicas do vendedor e historico de atendimento para PCs e hardware revisados.</p>
        </article>
        <article class="review-card">
          <div class="stars">★★★★★</div>
          <h3>Facebook Marketplace</h3>
          <p>Contato direto, retirada local em Vila Suzana e negociacao com fotos reais do produto.</p>
        </article>
        <article class="review-card">
          <div class="stars">★★★★★</div>
          <h3>Pos-venda</h3>
          <p>Garantia informada, suporte por WhatsApp e orientacao sobre upgrades antes da compra.</p>
        </article>
      </div>
    </section>

    <section class="section shell">
      <div class="brand-wall">
        <h2>Marcas confiaveis no ecossistema MobilyTech</h2>
        <div class="brand-row">
__BRAND_LOGOS__
        </div>
      </div>
    </section>

    <section class="section shell">
      <div class="section-head">
        <div>
          <h2>Regras claras antes da compra</h2>
          <p class="sub">Informacoes preventivas para manter a compra transparente e alinhada ao e-commerce brasileiro.</p>
        </div>
      </div>
      <div class="policy-grid">
        <article class="policy-card">
          <h3>Arrependimento online</h3>
          <p>Em compras online, o direito de arrependimento de 7 dias deve ser respeitado quando aplicavel.</p>
        </article>
        <article class="policy-card">
          <h3>Garantia MobilyTech</h3>
          <p>PCs revisados podem ter garantia comercial de 14 dias para defeitos preexistentes comprovados, sem cobrir danos por mau uso, alteracao indevida ou manipulacao apos a entrega.</p>
        </article>
        <article class="policy-card">
          <h3>MobilyTech Finds</h3>
          <p>A curadoria informa loja de origem, prazo, devolucao e suporte do marketplace para o cliente comprar com mais clareza.</p>
        </article>
      </div>
    </section>
  </main>

  <footer id="contato">
    <div class="shell footer-grid">
      <div class="brand-block">
        <div class="brand"><img src="./assets/mobilytech-logo.png" alt="" /><span>MobilyTech BR</span></div>
        <p>Vila Suzana, Sao Paulo, SP</p>
        <p>mobilytechbr@gmail.com</p>
        <p>WhatsApp: +55 (11) 95480-1967</p>
      </div>
      <div>
        <h3>Loja</h3>
        <a href="./fase2/ofertas.html">PC Gamer</a>
        <a href="./fase2/achados.html">MobilyTech Finds</a>
        <a href="./fase2/montagem.html">Montagem</a>
        <a href="./fase2/limpeza.html">Limpeza</a>
      </div>
      <div>
        <h3>Suporte</h3>
        <a href="https://wa.me/5511954801967">WhatsApp</a>
        <a href="mailto:mobilytechbr@gmail.com">E-mail</a>
        <a href="https://avaliacoes.olx.com.br/vendedor/859fd666-c047-4d6d-adac-374dd530d56c">Avaliacoes OLX</a>
      </div>
      <div>
        <h3>Empresa</h3>
        <a href="./fase2/avaliacoes.html">Avaliacoes</a>
        <a href="./index.html">Site original</a>
        <a href="./admin/index.html">Painel legado</a>
      </div>
      <div>
        <h3>Politicas</h3>
        <a href="./fase2/contato.html#politicas">Politica de trocas</a>
        <a href="./fase2/contato.html#entrega">Entrega</a>
        <a href="./fase2/contato.html#privacidade">Privacidade</a>
      </div>
    </div>
  </footer>

  <aside id="drawer" class="drawer" aria-live="polite"></aside>

  <script>
    const state = { products: [], swaps: [], addons: [], finalists: [], cart: [], heroIndex: 0 };
    const checkoutUrl = './index.html#cart';
    const money = value => new Intl.NumberFormat('pt-BR', { style: 'currency', currency: 'BRL' }).format(value || 0);
    const qs = selector => document.querySelector(selector);

    async function loadData() {
      const [products, swaps, addons, phase2] = await Promise.all([
        fetch('./data/products.json').then(r => r.json()),
        fetch('./data/swaps.json').then(r => r.json()),
        fetch('./data/addons.json').then(r => r.json()),
        fetch('./data/phase2-finalists.json').then(r => r.json())
      ]);
      state.products = products.filter(p => p.active !== false);
      state.swaps = swaps.filter(s => s.active !== false);
      state.addons = addons.filter(a => a.active !== false);
      state.finalists = phase2.finalists || [];
      renderProducts();
      renderFinalists();
      renderCreatives();
      startHero();
    }

    function productSpecs(product) {
      const specs = product.specs || {};
      return [specs.processor, specs.memory, specs.gpu, specs.storage, specs.powerSupply, specs.brand, specs.capacity]
        .filter(Boolean)
        .slice(0, 4)
        .join(' | ');
    }

    function renderProducts() {
      const term = qs('#searchInput').value.trim().toLowerCase();
      const grid = qs('#productGrid');
      grid.innerHTML = '';
      state.products
        .filter(product => !term || `${product.title} ${product.category} ${productSpecs(product)}`.toLowerCase().includes(term))
        .forEach(product => {
          const card = document.createElement('article');
          card.className = 'deal-card';
          card.innerHTML = `
            <div class="deal-img"><img src="${product.cutout || product.image}" alt="${product.title}"></div>
            <h3>${product.title}</h3>
            <p class="specs">${productSpecs(product) || product.badge || 'Catalogo MobilyTech'}</p>
            <div class="price">${money(product.price)}</div>
            <div class="deal-actions">
              <button class="ghost" type="button" data-config="${product.id}">Configurar</button>
              <button class="primary" type="button" data-add="${product.id}">Adicionar</button>
            </div>`;
          grid.appendChild(card);
        });
    }

    function renderFinalists() {
      const term = qs('#searchInput').value.trim().toLowerCase();
      const grid = qs('#finalistGrid');
      grid.innerHTML = '';
      state.finalists
        .filter(item => !term || `${item.title} ${item.niche} ${item.whySell}`.toLowerCase().includes(term))
        .forEach(item => {
          const market = item.marketplace || { name: 'Marketplace', logo: 'assets/mercado-livre-logo.svg', button: 'Ver oferta', class: 'market-ml' };
          const href = item.affiliateUrl || item.sourceUrl;
          const card = document.createElement('article');
          card.className = 'finalist-card';
          card.innerHTML = `
            <div class="affiliate-photo"><img src="${item.productImage}" alt="${item.title}"></div>
            <div class="market-row">
              <span class="badge">${item.confidence} confianca</span>
              <span class="market-badge"><img src="${market.logo}" alt="">${market.name}</span>
            </div>
            <h3>${item.title}</h3>
            <p>${item.whySell}</p>
            <p><strong>Preco:</strong> ${item.currentPrice}</p>
            <p class="model">${item.publicPartnerNote}</p>
            <span class="affiliate-status">${item.affiliateStatus}</span>
            <a class="market-button ${market.class}" href="${href}" target="_blank" rel="noreferrer">
              <img src="${market.logo}" alt="">${item.affiliateButton || market.button}
            </a>`;
          grid.appendChild(card);
        });
    }

    function renderCreatives() {
      const grid = qs('#creativeGrid');
      if (!grid) return;
      grid.innerHTML = '';
      state.finalists.forEach(item => {
        (item.creatives || []).forEach(creative => {
          const card = document.createElement('a');
          card.className = 'creative-card';
          card.href = creative.file;
          card.target = '_blank';
          card.rel = 'noreferrer';
          card.innerHTML = `<img src="${creative.file}" alt="Criativo ${item.title}"><span>${creative.selected ? 'Selecionado | ' : ''}${item.title} | ${creative.angle}</span>`;
          grid.appendChild(card);
        });
      });
    }

    function startHero() {
      const pcs = state.products.filter(p => p.category === 'pc' && (p.cutout || p.image));
      if (!pcs.length) return;
      const img = qs('#heroProduct');
      setInterval(() => {
        state.heroIndex = (state.heroIndex + 1) % pcs.length;
        img.style.opacity = '0';
        img.style.transform = 'translateY(10px) scale(.98)';
        setTimeout(() => {
          img.src = pcs[state.heroIndex].cutout || pcs[state.heroIndex].image;
          img.alt = pcs[state.heroIndex].title;
          img.style.opacity = '1';
          img.style.transform = 'translateY(0) scale(1)';
        }, 260);
      }, 4600);
    }

    function compatibleSwaps(product) {
      const specsText = Object.values(product.specs || {}).join(' ').toLowerCase();
      const specific = (product.swaps && product.swaps.processor ? product.swaps.processor : []).map((swap, index) => ({
        id: `${product.id}-processor-${index}`,
        label: swap.label,
        price: swap.price || 0
      }));
      const global = state.swaps.filter(swap => {
        const when = (swap.whenContains || []).every(token => specsText.includes(String(token).toLowerCase()));
        const exclude = (swap.excludeContains || []).some(token => specsText.includes(String(token).toLowerCase()));
        return (when || !swap.whenContains) && !exclude;
      });
      return [...specific, ...global];
    }

    function openConfig(productId) {
      const product = state.products.find(p => p.id === productId);
      if (!product) return;
      const drawer = qs('#drawer');
      const swaps = compatibleSwaps(product);
      let selected = [];
      let total = product.price || 0;
      const render = () => {
        total = (product.price || 0) + selected.reduce((sum, option) => sum + option.price, 0);
        drawer.innerHTML = `
          <button class="ghost" type="button" id="closeDrawer">Fechar</button>
          <h2>${product.title}</h2>
          <p class="specs">${productSpecs(product)}</p>
          <div class="price">${money(total)}</div>
          <h3>Trocas e adicionais</h3>
          ${[...swaps, ...state.addons].map(option => `
            <label class="option">
              <input type="checkbox" data-option="${option.id}" ${selected.some(s => s.id === option.id) ? 'checked' : ''}>
              <span>${option.label} <strong>${option.price >= 0 ? '+' : ''}${money(option.price)}</strong></span>
            </label>`).join('') || '<p class="specs">Nenhum adicional disponivel para este item.</p>'}
          <button class="primary" type="button" id="addConfigured">Adicionar configurado</button>
          <p class="sr-note">Finalize a compra pelo metodo disponivel no site ou chame a MobilyTech para combinar pagamento, frete e retirada.</p>`;
        drawer.querySelector('#closeDrawer').onclick = closeDrawer;
        drawer.querySelectorAll('[data-option]').forEach(input => {
          input.onchange = event => {
            const id = event.target.dataset.option;
            const option = [...swaps, ...state.addons].find(item => item.id === id);
            if (!option) return;
            selected = event.target.checked ? [...selected, option] : selected.filter(item => item.id !== id);
            render();
          };
        });
        drawer.querySelector('#addConfigured').onclick = () => {
          state.cart.push({
            id: `${product.id}-${Date.now()}`,
            productId: product.id,
            title: product.title,
            total,
            image: product.cutout || product.image,
            options: selected.map(s => s.label)
          });
          updateCart();
          openCart();
        };
      };
      drawer.classList.add('open');
      render();
    }

    function addProduct(productId) {
      const product = state.products.find(p => p.id === productId);
      if (!product) return;
      state.cart.push({
        id: `${product.id}-${Date.now()}`,
        productId: product.id,
        title: product.title,
        total: product.price || 0,
        image: product.cutout || product.image,
        options: []
      });
      updateCart();
      openCart();
    }

    function updateCart() {
      qs('#cartCount').textContent = state.cart.length;
    }

    function cartItemTemplate(item) {
      const options = item.options.length ? item.options.join(', ') : 'Sem adicionais';
      return `<article class="cart-line">
        <img src="${item.image || './assets/mobilytech-logo.png'}" alt="">
        <div>
          <strong>${item.title}</strong>
          <div class="cart-options">${options}</div>
          <div class="cart-line-bottom">
            <span>${money(item.total)}</span>
            <button class="cart-remove" type="button" data-remove-cart="${item.id}">Remover</button>
          </div>
        </div>
      </article>`;
    }

    function openCart() {
      const drawer = qs('#drawer');
      const total = state.cart.reduce((sum, item) => sum + item.total, 0);
      const shippingMessage = state.cart.length
        ? 'Frete Melhor Envio/retirada e formas de pagamento continuam no checkout original.'
        : 'Adicione um PC ou hardware para preparar o checkout.';
      drawer.innerHTML = `
        <div class="drawer-head">
          <div>
            <span class="drawer-kicker">Checkout MobilyTech</span>
            <h2>Seu carrinho</h2>
          </div>
          <button class="ghost drawer-close" type="button" id="closeDrawer">Fechar</button>
        </div>
        <div class="cart-list">
          ${state.cart.map(cartItemTemplate).join('') || '<p class="specs">Carrinho vazio.</p>'}
        </div>
        <div class="checkout-card">
          <div class="checkout-row"><span>Subtotal</span><strong>${money(total)}</strong></div>
          <input class="mini-field" id="cartCep" inputmode="numeric" placeholder="CEP para calcular frete no checkout">
          <div class="payment-chips" aria-label="Metodos preservados">
            <span>Wix Payments</span>
            <span>Mercado Pago</span>
            <span>Abacate Pay</span>
            <span>Pix QR Code</span>
            <span>Melhor Envio</span>
          </div>
          <div class="cart-actions">
            <a class="primary" href="${checkoutUrl}">Continuar para checkout seguro</a>
            <a class="ghost" href="https://wa.me/5511954801967?text=${encodeURIComponent('Quero finalizar um pedido MobilyTech pelo carrinho do site.')}">Falar no WhatsApp</a>
          </div>
          <p class="sr-note">${shippingMessage}</p>
        </div>`;
      drawer.classList.add('open');
      drawer.querySelector('#closeDrawer').onclick = closeDrawer;
      drawer.querySelectorAll('[data-remove-cart]').forEach(button => {
        button.onclick = event => {
          const id = event.currentTarget.dataset.removeCart;
          state.cart = state.cart.filter(item => item.id !== id);
          updateCart();
          openCart();
        };
      });
    }

    function closeDrawer() { qs('#drawer').classList.remove('open'); }

    document.addEventListener('click', event => {
      const configId = event.target.closest('[data-config]')?.dataset.config;
      const addId = event.target.closest('[data-add]')?.dataset.add;
      if (configId) openConfig(configId);
      if (addId) addProduct(addId);
    });
    qs('#cartButton').addEventListener('click', openCart);
    qs('#searchInput').addEventListener('input', () => { renderProducts(); renderFinalists(); });
    loadData().catch(error => {
      document.body.insertAdjacentHTML('afterbegin', `<pre style="color:#fff;background:#900;padding:12px">${error.message}</pre>`);
    });
  </script>
</body>
</html>
"""
    html_doc = html_doc.replace("__BRAND_LOGOS__", brand_logo_cards("./"))
    path = ROOT / "fase2-hibrida.html"
    path.write_text(html_doc, encoding="utf-8")
    return path


def brl(value: float | int | None) -> str:
    if value is None:
        return "Sob consulta"
    whole = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {whole}"


def page_nav(prefix: str = "../", active: str = "") -> str:
    items = [
        ("Inicio", f"{prefix}fase2-hibrida.html", "home"),
        ("Ofertas", "ofertas.html", "ofertas"),
        ("Monte seu PC", "montagem.html", "montagem"),
        ("Limpeza", "limpeza.html", "limpeza"),
        ("MobilyTech Finds", "achados.html", "achados"),
        ("Avaliacoes", "avaliacoes.html", "avaliacoes"),
        ("Contato", "contato.html", "contato"),
    ]
    return "\n".join(
        f'<a class="{"active" if key == active else ""}" href="{href}"><span class="nav-icon" aria-hidden="true"><svg viewBox="0 0 24 24">{NAV_ICONS.get(key, "")}</svg></span>{label}</a>'
        for label, href, key in items
    )


def subpage_css() -> str:
    return """    :root {
      color-scheme: dark;
      --bg: #02050a;
      --panel: rgba(7, 22, 34, .84);
      --line: rgba(121,247,255,.2);
      --cyan: #17d9ff;
      --green: #00ffc6;
      --text: #f6fbff;
      --muted: #a9bed0;
      font-family: Nunito, Inter, Segoe UI, Arial, sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      background:
        radial-gradient(circle at 18% 8%, rgba(23,217,255,.14), transparent 26rem),
        radial-gradient(circle at 82% 12%, rgba(0,255,198,.09), transparent 25rem),
        linear-gradient(180deg, #02050a 0%, #06101b 48%, #02050a 100%);
      color: var(--text);
      letter-spacing: 0;
      overflow-x: hidden;
    }
    a { color: inherit; text-decoration: none; }
    .shell { width: min(1440px, calc(100% - 36px)); margin: 0 auto; }
    .top-nav {
      position: sticky;
      top: 0;
      z-index: 20;
      background: rgba(2,7,13,.9);
      backdrop-filter: blur(18px);
      border-bottom: 1px solid var(--line);
    }
    .nav-row {
      min-height: 78px;
      display: grid;
      grid-template-columns: 230px 1fr;
      align-items: center;
      gap: 18px;
    }
    .brand { display: flex; align-items: center; gap: 10px; font-weight: 950; font-size: 20px; }
    .brand img { width: 40px; height: 40px; object-fit: contain; filter: drop-shadow(0 0 10px rgba(23,217,255,.42)); }
    .menu { display: flex; justify-content: flex-start; gap: 7px; overflow-x: auto; scrollbar-width: none; }
    .menu::-webkit-scrollbar { display: none; }
    .menu a {
      min-height: 36px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 6px;
      white-space: nowrap;
      font-size: 12px;
      font-weight: 950;
      color: #dff8ff;
      padding: 0 10px;
      border: 1px solid rgba(121,247,255,.16);
      border-radius: 999px;
      background: rgba(255,255,255,.045);
    }
    .menu a.active, .menu a:hover { color: var(--cyan); }
    .nav-icon {
      width: 16px;
      height: 16px;
      display: inline-grid;
      place-items: center;
      color: var(--cyan);
      flex: 0 0 auto;
    }
    .nav-icon svg {
      width: 16px;
      height: 16px;
      fill: none;
      stroke: currentColor;
      stroke-width: 1.9;
      stroke-linecap: round;
      stroke-linejoin: round;
    }
    .page-hero {
      margin: 28px auto 22px;
      min-height: 310px;
      display: grid;
      grid-template-columns: minmax(0, 1fr) 430px;
      gap: 24px;
      align-items: center;
      border-radius: 28px;
      padding: 42px 54px;
      border: 1px solid rgba(121,247,255,.2);
      background:
        linear-gradient(105deg, rgba(3,7,13,.96), rgba(4,32,45,.86), rgba(3,7,13,.72)),
        url("../assets/cleaning-neon-bg.png") center/cover;
      overflow: hidden;
      box-shadow: 0 22px 65px rgba(0,0,0,.28);
    }
    .eyebrow {
      display: inline-flex;
      padding: 7px 12px;
      border-radius: 999px;
      border: 1px solid rgba(23,217,255,.32);
      background: rgba(23,217,255,.12);
      color: #79f7ff;
      font-size: 12px;
      font-weight: 950;
      text-transform: uppercase;
    }
    h1 { margin: 16px 0 12px; font-size: clamp(36px, 5vw, 70px); line-height: .95; }
    h2 { margin: 0 0 16px; font-size: clamp(28px, 3vw, 44px); line-height: 1; }
    h3 { margin: 0; }
    p { color: #d6e8f3; font-weight: 760; line-height: 1.45; }
    .page-hero p { max-width: 760px; font-size: 17px; }
    .page-hero img { max-width: 100%; max-height: 280px; object-fit: contain; justify-self: end; filter: drop-shadow(0 18px 24px rgba(0,0,0,.34)); }
    .section { padding: 30px 0; }
    .grid { display: grid; gap: 18px; }
    .grid.products { grid-template-columns: repeat(4, minmax(0, 1fr)); }
    .grid.finalists { grid-template-columns: repeat(4, minmax(0, 1fr)); }
    .grid.two { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    .grid.three { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    .card {
      border-radius: 22px;
      background: var(--panel);
      border: 1px solid rgba(121,247,255,.18);
      box-shadow: 0 18px 48px rgba(0,0,0,.24);
      overflow: hidden;
      min-width: 0;
    }
    .product-card { padding: 16px; display: flex; flex-direction: column; min-height: 342px; }
    .product-card .imgbox { height: 166px; display: grid; place-items: center; border-radius: 18px; background: radial-gradient(circle, rgba(23,217,255,.2), rgba(0,0,0,.16) 58%); overflow: hidden; }
    .product-card img { max-width: 96%; max-height: 154px; object-fit: contain; }
    .product-card h3 { margin: 14px 0 8px; font-size: 17px; line-height: 1.18; }
    .specs { color: var(--muted); font-size: 12px; font-weight: 850; line-height: 1.4; }
    .price { color: var(--cyan); font-size: 25px; font-weight: 950; margin-top: auto; }
    .actions { display: flex; gap: 8px; margin-top: 12px; }
    .btn {
      display: inline-grid;
      place-items: center;
      min-height: 40px;
      border-radius: 999px;
      padding: 0 18px;
      background: linear-gradient(135deg, var(--cyan), var(--green));
      color: #041018;
      font-weight: 950;
      text-align: center;
      line-height: 1;
      text-decoration: none;
      white-space: nowrap;
      vertical-align: middle;
    }
    .btn.ghost { color: #79f7ff; background: transparent; border: 1px solid rgba(121,247,255,.42); }
    .actions .btn { flex: 1; font-size: 12px; }
    .text-card { padding: 24px; }
    .text-card h3 { color: var(--cyan); margin-bottom: 10px; font-size: 22px; }
    .text-card p { margin: 0 0 12px; }
    .service-form {
      padding: 24px;
      display: grid;
      gap: 14px;
    }
    .service-form h3 { color: var(--cyan); font-size: 24px; }
    .service-form p { margin: 0; }
    .field-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
    }
    .field-grid label {
      display: grid;
      gap: 6px;
      color: #dff7ff;
      font-size: 12px;
      font-weight: 950;
      text-transform: uppercase;
    }
    .field-grid label.full { grid-column: 1 / -1; }
    .field-grid input, .field-grid textarea {
      width: 100%;
      border: 1px solid rgba(121,247,255,.22);
      border-radius: 14px;
      background: rgba(0,0,0,.25);
      color: var(--text);
      outline: 0;
      padding: 12px 13px;
      font: inherit;
      font-weight: 850;
      text-transform: none;
    }
    .field-grid textarea { min-height: 94px; resize: vertical; }
    .form-actions { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    .form-actions .btn { min-width: 170px; min-height: 44px; }
    .form-note {
      color: var(--muted);
      font-size: 12px;
      font-weight: 850;
      line-height: 1.35;
    }
    .badge {
      display: inline-flex;
      align-self: flex-start;
      border-radius: 999px;
      padding: 5px 9px;
      background: rgba(23,217,255,.12);
      border: 1px solid rgba(23,217,255,.32);
      color: #79f7ff;
      font-size: 11px;
      font-weight: 950;
    }
    .finalist-card { padding: 14px; display: flex; flex-direction: column; gap: 10px; min-height: 388px; }
    .affiliate-photo {
      width: 100%;
      aspect-ratio: 1 / .78;
      display: grid;
      place-items: center;
      border-radius: 18px;
      overflow: hidden;
      background: linear-gradient(145deg, rgba(255,255,255,.98), rgba(219,241,247,.9));
      border: 1px solid rgba(121,247,255,.2);
      box-shadow: inset 0 0 0 1px rgba(255,255,255,.5), 0 16px 28px rgba(0,0,0,.18);
    }
    .affiliate-photo img { width: auto; height: auto; max-width: 92%; max-height: 88%; object-fit: contain; display: block; border-radius: 12px; }
    .market-row { display: flex; align-items: center; justify-content: space-between; gap: 8px; }
    .market-badge {
      display: inline-flex;
      align-items: center;
      gap: 7px;
      border-radius: 999px;
      padding: 5px 9px;
      color: #071018;
      background: linear-gradient(135deg, #ffe600, #fff2a8);
      font-size: 10.5px;
      font-weight: 950;
      box-shadow: 0 0 18px rgba(255,230,0,.15);
    }
    .market-badge img { width: 18px; height: 18px; object-fit: contain; }
    .affiliate-status { font-size: 10px; color: #9fb3bd; font-weight: 850; line-height: 1.25; min-height: 26px; }
    .market-button {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      width: 100%;
      min-height: 42px;
      margin-top: auto;
      border-radius: 999px;
      border: 1px solid rgba(255,255,255,.2);
      color: #071018;
      font-size: 12px;
      font-weight: 950;
      text-decoration: none;
      text-align: center;
      padding: 0 12px;
      box-shadow: 0 0 26px rgba(255,230,0,.16), inset 0 1px 0 rgba(255,255,255,.45);
      backdrop-filter: blur(12px);
    }
    .market-button img { width: 27px; height: 27px; object-fit: contain; flex: 0 0 auto; }
    .market-button.market-ml { background: linear-gradient(135deg, #fff159 0%, #ffe000 42%, #28a8ff 120%); }
    .finalist-card h3 { font-size: 15px; line-height: 1.18; }
    .finalist-card p { margin: 0; font-size: 12px; color: var(--muted); }
    .creative-grid { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 16px; }
    .creative-card { padding: 12px; }
    .creative-card img { width: 100%; aspect-ratio: 1 / 1; object-fit: cover; border-radius: 16px; display: block; }
    .creative-card span { display: block; margin-top: 9px; font-size: 12px; font-weight: 900; color: #dffbff; }
    .brand-wall {
      border-radius: 30px;
      padding: clamp(22px, 4vw, 42px);
      background:
        radial-gradient(circle at 20% 0%, rgba(23,217,255,.14), transparent 34%),
        linear-gradient(145deg, rgba(10,18,28,.94), rgba(3,8,14,.96));
      border: 1px solid rgba(121,247,255,.2);
      box-shadow: 0 22px 58px rgba(0,0,0,.32), inset 0 0 0 1px rgba(255,255,255,.035);
    }
    .brand-row {
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: nowrap;
      gap: clamp(16px, 2vw, 30px);
      overflow-x: auto;
      scrollbar-width: none;
    }
    .brand-row::-webkit-scrollbar { display: none; }
    .brand-pill {
      flex: 0 0 auto;
      min-width: 74px;
      display: grid;
      place-items: center;
      padding: 0;
      border: 0;
      border-radius: 0;
      background: transparent;
      box-shadow: none;
      overflow: visible;
    }
    .brand-pill:nth-last-child(2), .brand-pill:last-child { grid-column: auto; }
    .brand-pill img { display: block; width: auto; height: 32px; max-width: 112px; max-height: 32px; object-fit: contain; filter: drop-shadow(0 0 10px rgba(23,217,255,.22)); }
    footer { margin-top: 40px; padding: 38px 0; border-top: 1px solid rgba(121,247,255,.16); background: #03070d; }
    .footer-grid { display: grid; grid-template-columns: 1.5fr repeat(4, 1fr); gap: 22px; }
    footer h3 { margin: 0 0 12px; font-size: 16px; }
    footer a, footer p { display: block; color: #c3d4df; font-weight: 800; font-size: 14px; margin: 8px 0; }
    @media (max-width: 1120px) {
      .page-hero { grid-template-columns: 1fr 330px; padding: 36px; }
      .grid.products { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .grid.finalists { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .brand-row { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    }
    @media (max-width: 760px) {
      .shell { width: min(100% - 22px, 720px); }
      .nav-row { grid-template-columns: 1fr; gap: 10px; padding: 12px 0; }
      .brand { font-size: 16px; }
      .brand img { width: 34px; height: 34px; }
      .menu { justify-content: start; gap: 7px; padding-right: 20px; mask-image: linear-gradient(90deg, #000 0%, #000 calc(100% - 24px), transparent 100%); }
      .menu a { min-height: 31px; font-size: 10.5px; padding: 0 10px; }
      .nav-icon, .nav-icon svg { width: 14px; height: 14px; }
      .page-hero { grid-template-columns: 1fr; min-height: auto; padding: 24px 18px; border-radius: 20px; gap: 14px; }
      .page-hero img { max-height: 150px; justify-self: center; }
      h1 { font-size: 30px; line-height: 1.02; }
      .page-hero p { font-size: 13px; }
      .grid.products, .grid.finalists, .creative-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px; }
      .grid.two, .grid.three { grid-template-columns: 1fr; }
      .field-grid { grid-template-columns: 1fr; }
      .product-card { min-height: 270px; padding: 10px; border-radius: 16px; }
      .product-card .imgbox { height: 112px; border-radius: 14px; }
      .product-card img { max-height: 104px; }
      .product-card h3 { font-size: 12px; }
      .specs, .finalist-card p { font-size: 10.5px; }
      .price { font-size: 18px; }
      .actions { flex-direction: column; }
      .actions .btn { min-height: 32px; font-size: 10.5px; }
      .form-actions .btn { width: 100%; min-width: 0; }
      .finalist-card { min-height: 318px; padding: 10px; border-radius: 16px; }
      .affiliate-photo { border-radius: 13px; }
      .market-row { align-items: flex-start; flex-direction: column; gap: 6px; }
      .market-button { min-height: 34px; font-size: 10px; gap: 5px; white-space: normal; line-height: 1.12; padding: 5px 8px; }
      .market-button img { width: 21px; height: 21px; }
      .affiliate-status { font-size: 9px; min-height: 0; }
      .finalist-card h3 { font-size: 12px; }
      .creative-card span { font-size: 10.5px; line-height: 1.25; }
      .brand-wall { padding: 20px 12px; }
      .brand-row { display: flex; flex-wrap: nowrap; gap: 18px; overflow-x: auto; }
      .brand-pill, .brand-pill:nth-last-child(2), .brand-pill:last-child { grid-column: auto; }
      .brand-pill { min-width: 78px; }
      .brand-pill img { height: 27px; max-width: 92px; max-height: 27px; }
      .footer-grid { grid-template-columns: 1fr 1fr; }
      footer .brand-block { grid-column: 1 / -1; }
    }
"""


def product_specs(product: dict) -> str:
    specs = product.get("specs") or {}
    values = [
        specs.get("processor"),
        specs.get("memory"),
        specs.get("gpu"),
        specs.get("storage"),
        specs.get("powerSupply"),
        specs.get("brand"),
        specs.get("capacity"),
    ]
    return " | ".join(str(value) for value in values if value)[:170]


def product_card(product: dict, prefix: str = "../") -> str:
    image = product.get("cutout") or product.get("image") or ""
    if image.startswith("./"):
        image = prefix + image[2:]
    title = html.escape(product.get("title", "Produto MobilyTech"))
    specs = html.escape(product_specs(product) or product.get("badge", "Catalogo MobilyTech"))
    return f"""      <article class="card product-card">
        <div class="imgbox"><img src="{image}" alt="{title}"></div>
        <h3>{title}</h3>
        <p class="specs">{specs}</p>
        <div class="price">{brl(product.get("price"))}</div>
        <div class="actions">
          <a class="btn ghost" href="../fase2-hibrida.html#ofertas">Configurar</a>
          <a class="btn" href="https://wa.me/5511954801967?text=Tenho%20interesse%20em%20{quote(product.get("title", "produto MobilyTech"))}">Consultar</a>
        </div>
      </article>"""


def finalist_card(item: dict) -> str:
    market = item["marketplace"]
    product_image = item["productImage"][2:] if item["productImage"].startswith("./") else item["productImage"]
    market_logo = market["logo"]
    return f"""      <article class="card finalist-card">
        <div class="affiliate-photo"><img src="../{html.escape(product_image)}" alt="{html.escape(item["title"])}"></div>
        <div class="market-row">
          <span class="badge">{html.escape(item["confidence"])} confianca</span>
          <span class="market-badge"><img src="../{html.escape(market_logo)}" alt="">{html.escape(market["name"])}</span>
        </div>
        <h3>{html.escape(item["title"])}</h3>
        <p>{html.escape(item["whySell"])}</p>
        <p><strong>Preco:</strong> {html.escape(item["currentPrice"])}</p>
        <p>{html.escape(item["publicPartnerNote"])}</p>
        <span class="affiliate-status">{html.escape(item["affiliateStatus"])}</span>
        <a class="market-button {html.escape(market["class"])}" href="{html.escape(item["affiliateUrl"])}" target="_blank" rel="noreferrer">
          <img src="../{html.escape(market_logo)}" alt="">{html.escape(item["affiliateButton"])}
        </a>
      </article>"""


def shared_footer(prefix: str = "../") -> str:
    return f"""  <footer id="contato">
    <div class="shell footer-grid">
      <div class="brand-block">
        <div class="brand"><img src="{prefix}assets/mobilytech-logo.png" alt=""><span>MobilyTech BR</span></div>
        <p>Vila Suzana, Sao Paulo, SP</p>
        <p>mobilytechbr@gmail.com</p>
        <p>WhatsApp: +55 (11) 95480-1967</p>
      </div>
      <div>
        <h3>Loja</h3>
        <a href="ofertas.html">PC Gamer</a>
        <a href="achados.html">MobilyTech Finds</a>
        <a href="montagem.html">Montagem</a>
        <a href="limpeza.html">Limpeza</a>
      </div>
      <div>
        <h3>Suporte</h3>
        <a href="contato.html">Contato</a>
        <a href="https://wa.me/5511954801967">WhatsApp</a>
        <a href="mailto:mobilytechbr@gmail.com">E-mail</a>
      </div>
      <div>
        <h3>Empresa</h3>
        <a href="avaliacoes.html">Avaliacoes</a>
        <a href="../fase2-hibrida.html">Pagina principal</a>
        <a href="{prefix}admin/index.html">Painel legado</a>
      </div>
      <div>
        <h3>Politicas</h3>
        <a href="contato.html#politicas">Politica de trocas</a>
        <a href="contato.html#entrega">Entrega</a>
        <a href="contato.html#privacidade">Privacidade</a>
      </div>
    </div>
  </footer>"""


def subpage_html(active: str, title: str, subtitle: str, hero_image: str, content: str) -> str:
    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{html.escape(title)} | MobilyTech BR</title>
  <link rel="icon" href="../assets/favicon.png" />
  <style>
{subpage_css()}
  </style>
</head>
<body>
  <nav class="top-nav">
    <div class="shell nav-row">
      <a class="brand" href="../fase2-hibrida.html"><img src="../assets/mobilytech-logo.png" alt=""><span>MobilyTech BR</span></a>
      <div class="menu" aria-label="Navegacao principal">
{page_nav("../", active)}
      </div>
    </div>
  </nav>
  <header class="shell page-hero">
    <div>
      <span class="eyebrow">MobilyTech BR</span>
      <h1>{html.escape(title)}</h1>
      <p>{html.escape(subtitle)}</p>
    </div>
    <img src="{hero_image}" alt="">
  </header>
{content}
{shared_footer("../")}
  <script>
    (() => {{
      const form = document.querySelector('#cleaningScheduleForm');
      if (!form) return;
      form.addEventListener('submit', event => {{
        event.preventDefault();
        const data = new FormData(form);
        const name = String(data.get('name') || '').trim();
        const phone = String(data.get('phone') || '').trim();
        const email = String(data.get('email') || '').trim();
        const notes = String(data.get('notes') || '').trim();
        const message = [
          'Quero agendar uma limpeza de PC pela MobilyTech BR.',
          '',
          'Nome: ' + name,
          'Numero: ' + phone,
          'E-mail: ' + email,
          notes ? 'Observacoes: ' + notes : ''
        ].filter(Boolean).join('\\n');
        window.location.href = 'https://wa.me/5511954801967?text=' + encodeURIComponent(message);
      }});
    }})();
  </script>
</body>
</html>
"""


def write_subpages() -> list[Path]:
    out_dir = ROOT / "fase2"
    out_dir.mkdir(parents=True, exist_ok=True)
    products = [p for p in json.loads((DATA_DIR / "products.json").read_text(encoding="utf-8")) if p.get("active") is not False]
    pc_products = [p for p in products if p.get("category") == "pc"]
    hardware_products = [p for p in products if p.get("category") != "pc"]
    product_grid = "\n".join(product_card(product) for product in products)
    pc_grid = "\n".join(product_card(product) for product in pc_products)
    hardware_grid = "\n".join(product_card(product) for product in hardware_products)
    finalist_grid = "\n".join(finalist_card(item) for item in FINALISTS)
    creative_grid = "\n".join(
        f"""      <a class="card creative-card" href="../{creative['file'][2:]}" target="_blank" rel="noreferrer">
        <img src="../{creative['file'][2:]}" alt="Criativo {html.escape(item['title'])}">
        <span>{'Selecionado | ' if creative.get('selected') else ''}{html.escape(item['title'])} | {html.escape(creative['angle'])}</span>
      </a>"""
        for item in FINALISTS
        for creative in item.get("creatives", [])
    )
    brand_wall = f"""  <section class="section shell">
    <div class="brand-wall">
      <h2>Marcas confiaveis no ecossistema MobilyTech</h2>
      <div class="brand-row">
{brand_logo_cards("../")}
      </div>
    </div>
  </section>"""

    pages = {
        "index.html": subpage_html(
            "home",
            "MobilyTech BR",
            "Navegue pelas paginas principais da loja: PCs, montagem, limpeza, avaliacoes, contato e MobilyTech Finds.",
            "../assets/generated/pcryzen-5-3600-cutout.png",
            """  <section class="section shell">
    <div class="grid two">
      <a class="card text-card" href="ofertas.html"><h3>Ofertas e PCs</h3><p>Catalogo separado de PCs, SSDs, fonte e hardware real.</p></a>
      <a class="card text-card" href="achados.html"><h3>MobilyTech Finds</h3><p>Produtos tech selecionados para complementar setups, upgrades e manutencao.</p></a>
      <a class="card text-card" href="montagem.html"><h3>Monte seu PC</h3><p>Pagina propria para montagem sob orcamento.</p></a>
      <a class="card text-card" href="limpeza.html"><h3>Limpeza</h3><p>Pagina propria para limpeza e relatorio.</p></a>
    </div>
  </section>""" + brand_wall,
        ),
        "ofertas.html": subpage_html(
            "ofertas",
            "Ofertas e PCs revisados",
            "Catalogo separado para os PCs e hardwares reais da MobilyTech, mantendo o visual da home principal.",
            "../assets/generated/pc-gamer-i5-gt610-cutout.png",
            f"""  <section id="pcs" class="section shell">
    <h2>PC Gamer</h2>
    <div class="grid products">
{pc_grid}
    </div>
  </section>
  <section class="section shell">
    <h2>Hardware e pecas</h2>
    <div class="grid products">
{hardware_grid}
    </div>
  </section>""",
        ),
        "achados.html": subpage_html(
            "achados",
            "MobilyTech Finds",
            "Produtos tech selecionados para complementar setups, upgrades e manutencao, com compra segura em marketplaces.",
            "../assets/mobilytech-character-cutout.png",
            f"""  <section class="section shell">
    <h2>Produtos selecionados</h2>
    <div class="grid finalists">
{finalist_grid}
    </div>
  </section>
  <section class="section shell">
    <div class="card text-card">
      <h3>Curadoria MobilyTech</h3>
      <p>Esta selecao prioriza acessorios uteis para quem compra, monta ou cuida de PCs: armazenamento, organizacao, limpeza, conectividade e setup gamer.</p>
    </div>
  </section>""" + brand_wall,
        ),
        "montagem.html": subpage_html(
            "montagem",
            "Monte seu PC",
            "Pagina dedicada para montagem sob orcamento, com configuracao orientada por uso, jogos, limite de preco e disponibilidade de pecas.",
            "../assets/assembly-pc-build-cutout.png",
            """  <section class="section shell">
    <div class="grid two">
      <article class="card text-card"><h3>Orcamento personalizado</h3><p>O cliente informa objetivo, jogos, programas, limite de preco e preferencia por novo/usado. A MobilyTech monta uma proposta coerente.</p><a class="btn" href="https://wa.me/5511954801967?text=Quero%20montar%20um%20PC%20com%20a%20MobilyTech%20BR">Chamar no WhatsApp</a></article>
      <article class="card text-card"><h3>Garantia e transparencia</h3><p>Antes de fechar, a proposta precisa deixar claro pecas, estado, garantia, prazo e condicoes de suporte.</p><a class="btn ghost" href="contato.html#politicas">Ver regras</a></article>
    </div>
  </section>""" + brand_wall,
        ),
        "limpeza.html": subpage_html(
            "limpeza",
            "Limpeza de PCs",
            "Pagina dedicada para limpeza, relatorio visual e orientacao de manutencao preventiva.",
            "../assets/pc-cleaning-service-cutout.png",
            """  <section class="section shell">
    <div class="grid two">
      <article class="card text-card"><h3>Limpeza com relatorio</h3><p>Antes/depois, cuidado com poeira, organizacao visual e registro para o cliente acompanhar o servico.</p><a class="btn" href="#agendamento">Agendar limpeza</a></article>
      <article class="card text-card"><h3>Produtos relacionados</h3><p>Kits de limpeza e acessorios da curadoria MobilyTech complementam o servico principal sem misturar atendimento com venda.</p><a class="btn ghost" href="achados.html">Ver MobilyTech Finds</a></article>
    </div>
  </section>
  <section id="agendamento" class="section shell">
    <form class="card service-form" id="cleaningScheduleForm">
      <div>
        <span class="eyebrow">Agendamento</span>
        <h3>Solicitar limpeza de PC</h3>
        <p>Preencha seus dados para abrir o atendimento pelo WhatsApp com a mensagem pronta.</p>
      </div>
      <div class="field-grid">
        <label>Nome
          <input name="name" autocomplete="name" placeholder="Seu nome" required>
        </label>
        <label>Numero
          <input name="phone" autocomplete="tel" inputmode="tel" placeholder="(DDD) 9XXXX-XXXX" required>
        </label>
        <label class="full">E-mail
          <input name="email" type="email" autocomplete="email" placeholder="seuemail@email.com" required>
        </label>
        <label class="full">Observacoes
          <textarea name="notes" placeholder="Descreva o PC, problema de temperatura, poeira, barulho ou urgencia."></textarea>
        </label>
      </div>
      <div class="form-actions">
        <button class="btn" type="submit">Agende ja</button>
        <span class="form-note">O envio abre o WhatsApp da MobilyTech com seus dados preenchidos para confirmar horario, valor e retirada/entrega.</span>
      </div>
    </form>
  </section>""" + brand_wall,
        ),
        "avaliacoes.html": subpage_html(
            "avaliacoes",
            "Avaliacoes e confianca",
            "Pagina dedicada para prova social, canais reais e marcas presentes no ecossistema MobilyTech.",
            "../assets/mobilytech-character-cutout.png",
            """  <section class="section shell">
    <div class="grid three reviews-grid">
      <article class="card text-card"><h3>OLX</h3><p>Avaliacoes publicas e historico de vendedor para PCs e hardware revisados.</p><a class="btn ghost" href="https://avaliacoes.olx.com.br/vendedor/859fd666-c047-4d6d-adac-374dd530d56c">Ver OLX</a></article>
      <article class="card text-card"><h3>Facebook Marketplace</h3><p>Contato direto, retirada local e negociacao com fotos reais dos produtos.</p></article>
      <article class="card text-card"><h3>Pos-venda</h3><p>Suporte por WhatsApp, garantia informada e orientacao de upgrades antes da compra.</p></article>
    </div>
  </section>""" + brand_wall,
        ),
        "contato.html": subpage_html(
            "contato",
            "Contato, entrega e politicas",
            "Pagina dedicada para suporte, retirada local, politicas e pontos legais antes da compra.",
            "../assets/mobilytech-logo.png",
            """  <section class="section shell">
    <div class="grid two">
      <article class="card text-card"><h3>Contato</h3><p>E-mail: mobilytechbr@gmail.com<br>WhatsApp: +55 (11) 95480-1967<br>Retirada: Vila Suzana, Sao Paulo, SP</p><a class="btn" href="https://wa.me/5511954801967">Chamar no WhatsApp</a></article>
      <article id="entrega" class="card text-card"><h3>Entrega e retirada</h3><p>Frete e retirada precisam ser confirmados no checkout ou no atendimento, usando CEP, seguro e disponibilidade real.</p></article>
      <article id="politicas" class="card text-card"><h3>Trocas e garantia</h3><p>PCs revisados podem ter garantia comercial de 14 dias para defeitos preexistentes comprovados, sem cobrir dano por mau uso ou alteracao indevida apos entrega.</p></article>
      <article id="privacidade" class="card text-card"><h3>Privacidade</h3><p>Dados de contato, endereco e pagamento devem ser tratados apenas para atendimento, entrega, pos-venda e obrigacoes legais.</p></article>
    </div>
  </section>""",
        ),
    }
    written = []
    for filename, page in pages.items():
        path = out_dir / filename
        path.write_text(page, encoding="utf-8")
        written.append(path)
    return written


def write_csvs() -> tuple[Path, Path]:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    finalists_csv = DOCS_DIR / "phase2-finalists-2026-06-13.csv"
    creative_csv = DOCS_DIR / "phase2-creatives-2026-06-13.csv"
    fields = [
        "id",
        "title",
        "niche",
        "platform",
        "currentPrice",
        "shipping",
        "delivery",
        "sellerReputation",
        "reviews",
        "returnPolicy",
        "operationModel",
        "whySell",
        "costTarget",
        "sellTarget",
        "margin",
        "confidence",
        "risk",
        "sourceUrl",
        "googleTrends",
        "metaAdsLibrary",
        "tikTokCreativeCenter",
        "mercadoLivreTendencias",
    ]
    with finalists_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";")
        writer.writeheader()
        for item in FINALISTS:
            row = {key: item.get(key, "") for key in fields}
            for key in ("googleTrends", "metaAdsLibrary", "tikTokCreativeCenter", "mercadoLivreTendencias"):
                row[key] = item["researchLinks"].get(key, "")
            writer.writerow(row)
    with creative_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["productId", "product", "variant", "angle", "file", "status"], delimiter=";")
        writer.writeheader()
        for item in FINALISTS:
            for creative in item["creatives"]:
                writer.writerow(
                    {
                        "productId": item["id"],
                        "product": item["title"],
                        "variant": creative["variant"],
                        "angle": creative["angle"],
                        "file": creative["file"],
                        "status": creative["status"],
                    }
                )
    return finalists_csv, creative_csv


def write_workbook() -> Path:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Finalistas Fase 2"
    headers = [
        "Produto",
        "Nicho",
        "Modelo operacional",
        "Preco atual",
        "Frete",
        "Prazo/logistica",
        "Reputacao",
        "Avaliacoes",
        "Devolucao",
        "Margem possivel",
        "Confianca",
        "Risco",
        "Link principal",
        "Google Trends",
        "Meta Ads Library",
        "TikTok Creative Center",
        "ML Tendencias",
    ]
    ws.append(headers)
    for item in FINALISTS:
        ws.append(
            [
                item["title"],
                item["niche"],
                item["operationModel"],
                item["currentPrice"],
                item["shipping"],
                item["delivery"],
                item["sellerReputation"],
                item["reviews"],
                item["returnPolicy"],
                item["margin"],
                item["confidence"],
                item["risk"],
                item["sourceUrl"],
                item["researchLinks"]["googleTrends"],
                item["researchLinks"]["metaAdsLibrary"],
                item["researchLinks"]["tikTokCreativeCenter"],
                item["researchLinks"]["mercadoLivreTendencias"],
            ]
        )

    ws2 = wb.create_sheet("Criativos")
    ws2.append(["Produto", "Variante", "Angulo", "Arquivo", "Status"])
    for item in FINALISTS:
        for creative in item["creatives"]:
            ws2.append([item["title"], creative["variant"], creative["angle"], creative["file"], creative["status"]])

    ws3 = wb.create_sheet("Execucao")
    exec_rows = [
        ["Gerado em", GENERATED_AT],
        ["Backup usado", str(BACKUP_PATH)],
        ["Preview local", "http://127.0.0.1:4173/fase2-hibrida.html"],
        ["Arquivo HTML", str(ROOT / "fase2-hibrida.html")],
        ["JSON", str(DATA_DIR / "phase2-finalists.json")],
        ["Wix Premium site ID", WIX_PREMIUM_SITE_ID],
        ["Dominio Wix", WIX_DOMAIN],
        ["Status anuncios pagos", "Bloqueados ate aprovacao manual"],
        ["Pendencia legal", "Direito de arrependimento de 7 dias deve permanecer para compra online quando aplicavel"],
    ]
    for row in exec_rows:
        ws3.append(row)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="062131")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="1F5A6A")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_length = min(max((len(str(cell.value or "")) for cell in column), default=10), 58)
            sheet.column_dimensions[letter].width = max(12, max_length + 2)

    docs_xlsx = DOCS_DIR / "MobilyTech_Fase2_Finalistas_Validacao_Criativos_2026-06-13.xlsx"
    out_xlsx = OUTPUT_DIR / docs_xlsx.name
    wb.save(docs_xlsx)
    wb.save(out_xlsx)
    return docs_xlsx


def write_report() -> Path:
    lines = [
        "# MobilyTech BR - Fase 2 Hibrida",
        "",
        f"Gerado em: {GENERATED_AT}",
        "",
        "## Objetivo",
        "",
        "Criar uma versao alternativa no Vercel, sem alterar a home principal, usando a base visual e funcional da MobilyTech BR, mas com estrutura inspirada principalmente na iBUYPOWER e parcialmente na KaBuM. A pagina tambem prepara a ponte para Wix Stores, marketing e produtos de afiliado/dropshipping depois de aprovacao.",
        "",
        "## Arquivos principais",
        "",
        f"- Backup: `{BACKUP_PATH}`",
        "- Preview local: `http://127.0.0.1:4173/fase2-hibrida.html`",
        "- HTML: `fase2-hibrida.html`",
        "- Dados: `data/phase2-finalists.json`",
        "- Criativos: `assets/phase2-creatives/*.jpg`",
        "- Planilha: `docs/MobilyTech_Fase2_Finalistas_Validacao_Criativos_2026-06-13.xlsx`",
        "",
        "## O que foi aplicado da referencia iBUYPOWER/KaBuM",
        "",
        "- Barra superior promocional, navegacao por categorias e busca.",
        "- Hero grande com PC real da MobilyTech, chamada de campanha e visual de loja gamer.",
        "- Cards grandes para PCs prontos, montagem personalizada e limpeza de PCs.",
        "- Bloco de avaliacoes no estilo cards brancos, adaptado para OLX/Facebook/pos-venda.",
        "- Rodape com Loja, Suporte, Empresa e Legal.",
        "- Fileira de marcas confiaveis com proporcao controlada para nao cortar logos/textos.",
        "- Grid mobile reduzido para caber mais itens por linha e evitar botoes gigantes.",
        "",
        "## Finalistas escolhidos",
        "",
    ]
    for index, item in enumerate(FINALISTS, start=1):
        lines.extend(
            [
                f"### {index}. {item['title']}",
                "",
                f"- Nicho: {item['niche']}",
                f"- Modelo: {item['operationModel']}",
                f"- Preco/frete: {item['currentPrice']} | {item['shipping']}",
                f"- Reputacao: {item['sellerReputation']}",
                f"- Por que vender: {item['whySell']}",
                f"- Margem: {item['margin']}",
                f"- Risco: {item['risk']}",
                f"- Link: {item['sourceUrl']}",
                f"- Pesquisa: [Google Trends]({item['researchLinks']['googleTrends']}) | [Meta Ads Library]({item['researchLinks']['metaAdsLibrary']}) | [TikTok Creative Center]({item['researchLinks']['tikTokCreativeCenter']}) | [ML Tendencias]({item['researchLinks']['mercadoLivreTendencias']})",
                "",
            ]
        )
    lines.extend(
        [
            "## Observacoes de conformidade",
            "",
            "- Para compras online, o direito de arrependimento de 7 dias nao deve ser removido quando aplicavel. Fontes consultadas: gov.br/MJ e CDC/Planalto.",
            "- A garantia comercial de 14 dias para PCs revisados pode ser exibida como garantia adicional/contratual, mas nao deve excluir direitos legais por vicio/defeito.",
            "- Produtos de afiliado/dropshipping precisam deixar prazo, origem, devolucao, suporte e responsavel pela entrega claros antes de ir ao ar.",
            "",
            "## Pendencias antes de campanha paga",
            "",
            "- Confirmar no link final de cada produto: preco atual, prazo para CEP de destino, vendedor, reputacao e devolucao.",
            "- Validar manualmente Google Trends, Meta Ads Library, TikTok Creative Center e Mercado Livre Tendencias nos links preparados.",
            "- Aprovar 2 criativos por produto e definir orcamento pequeno de teste.",
            "- So depois subir teste e medir CPC, CTR, conversao, margem e suporte.",
        ]
    )
    path = DOCS_DIR / "phase2-hibrida-execution-2026-06-13.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def main() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_brand_logos()
    write_creatives()
    json_path = write_phase2_json()
    page_path = write_page()
    subpages = write_subpages()
    finalists_csv, creative_csv = write_csvs()
    xlsx_path = write_workbook()
    report_path = write_report()
    print(json.dumps(
        {
            "json": str(json_path),
            "page": str(page_path),
            "subpages": [str(path) for path in subpages],
            "finalistsCsv": str(finalists_csv),
            "creativeCsv": str(creative_csv),
            "xlsx": str(xlsx_path),
            "report": str(report_path),
            "creatives": len(list(CREATIVE_DIR.glob("*.jpg"))),
            "brandLogos": len(list((ROOT / "assets" / "brand-tiles").glob("*.jpg"))),
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
